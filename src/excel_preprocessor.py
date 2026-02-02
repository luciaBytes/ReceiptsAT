"""
Excel Preprocessor for Landlord Monthly Payment Files.

This module handles parsing of landlord Excel files with yearly tenant payment data
and generates receipt records for a selected month. Implements date-intelligent 
processing with cross-column payment detection for late/early payments.

Key Features:
- Scans all 12 month columns for payments matching selected month
- Detects cross-column payments (late/early scenarios)
- Validates Excel structure and data before processing
- Generates alerts for anomalous payment patterns
- Produces CSV-compatible receipt data
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date
from dataclasses import dataclass, field
from typing import List, Tuple, Optional
from pathlib import Path

from utils.logger import get_logger

try:
    from date_calculator import RentPeriodCalculator
except ImportError:
    from src.date_calculator import RentPeriodCalculator

logger = get_logger(__name__)


# Month name mapping (column headers to month numbers)
MONTH_COLUMNS = {
    'Jan': 1, 'January': 1,
    'Feb': 2, 'February': 2,
    'Mar': 3, 'March': 3,
    'Apr': 4, 'April': 4,
    'May': 5,
    'Jun': 6, 'June': 6,
    'Jul': 7, 'July': 7,
    'Aug': 8, 'August': 8,
    'Sep': 9, 'September': 9,
    'Oct': 10, 'October': 10,
    'Nov': 11, 'November': 11,
    'Dec': 12, 'December': 12
}


@dataclass
class TenantData:
    """
    Represents a tenant row from the Excel file.
    
    Attributes:
        contract_number: Contract identifier
        name: Tenant name (informational)
        rent: Monthly rent amount
        rent_deposit: Months paid in advance (typically 1)
        deposit_month_offset: Month offset from "Mês Caução" column (for from/to dates)
        months_late: Months behind on payment (0 if on time)
        paid_current_month: True if paying for current month instead of future
        row_number: Excel row number (for error reporting)
    """
    contract_number: str
    name: str
    rent: float
    rent_deposit: int
    deposit_month_offset: int
    months_late: int
    paid_current_month: bool
    row_number: int


@dataclass
class PaymentInfo:
    """
    Represents a single payment found in a month column.
    
    Attributes:
        tenant: Reference to TenantData
        payment_date: Calculated date of payment
        payment_column: Column where payment was found (month name)
        payment_day: Day of month payment was made
    """
    tenant: TenantData
    payment_date: date
    payment_column: str
    payment_day: int


@dataclass
class ProcessingAlert:
    """
    Alert for cross-column payment detection.
    
    Attributes:
        contract_number: Contract with anomalous payment
        payment_date: Date payment was made
        payment_column: Column where payment was found
        expected_column: Column where payment was expected
        rent_period_from: Start of rent period
        rent_period_to: End of rent period
        reason: Description of why alert was generated
    """
    contract_number: str
    payment_date: date
    payment_column: str
    expected_column: str
    rent_period_from: date
    rent_period_to: date
    reason: str


@dataclass
class ReceiptData:
    """
    Receipt record ready for CSV export.
    
    Attributes:
        contract_id: Contract identifier
        from_date: Start of rent period
        to_date: End of rent period
        payment_date: Date payment was made
        receipt_type: Type of receipt (typically "rent")
        value: Payment amount
        rent_deposit: Months paid in advance (from TenantData)
        months_late: Months behind on payment (from TenantData)
    """
    contract_id: str
    from_date: date
    to_date: date
    payment_date: date
    receipt_type: str
    value: float
    rent_deposit: int = 0
    months_late: int = 0


class LandlordExcelProcessor:
    """
    Processes landlord Excel files to generate receipt records.
    
    This class handles parsing, validation, and transformation of Excel files
    containing yearly tenant payment data into receipt records for a specific month.
    """
    
    def __init__(self):
        """Initialize the processor."""
        self.validation_errors: List[str] = []
        self.processing_alerts: List[ProcessingAlert] = []
    
    def parse_excel(
        self,
        file_path: str,
        selected_month: int,
        selected_year: int,
        sheet_name: Optional[str] = None
    ) -> Tuple[List[ReceiptData], List[ProcessingAlert]]:
        """
        Parse Excel file and generate receipts for selected month.
        
        Args:
            file_path: Path to Excel file
            selected_month: Month to process (1-12)
            selected_year: Year to process
            sheet_name: Name of sheet/tab to read (if None, uses first sheet)
        
        Returns:
            Tuple of (receipt_list, alert_list)
        
        Raises:
            ValueError: If file cannot be parsed or validation fails
            FileNotFoundError: If file doesn't exist
        """
        logger.info(f"Smart Import: Processing month {selected_month}/{selected_year}")
        logger.info(f"Source file: {file_path}")
        
        # Reset state
        self.validation_errors = []
        self.processing_alerts = []
        
        # Validate file exists
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            # Use specified sheet or active sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {', '.join(workbook.sheetnames)}")
                worksheet = workbook[sheet_name]
                logger.info(f"Using sheet: {sheet_name}")
            else:
                worksheet = workbook.active
                logger.info(f"Using active sheet: {worksheet.title}")
        except ValueError:
            raise  # Re-raise sheet not found errors
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise ValueError(f"Cannot open Excel file: {e}")
        
        # Validate structure
        self._validate_excel_structure(worksheet)
        
        if self.validation_errors:
            error_msg = "\n".join(self.validation_errors)
            logger.error(f"Validation failed:\n{error_msg}")
            raise ValueError(f"Excel validation failed:\n\n{error_msg}")
        
        # Parse tenants
        tenants = self._parse_tenants(worksheet)
        logger.info(f"Loaded {len(tenants)} tenant records from Excel")
        
        # Get column map for reading payment dates
        header_row = [cell.value for cell in worksheet[1]]
        col_map = self._build_column_map(header_row)
        
        # Prepare receipt data records (NOT submitting to portal)
        receipts = self._prepare_receipt_records(tenants, selected_month, selected_year, worksheet, col_map)
        logger.info(f"Prepared {len(receipts)} receipt data records, {len(self.processing_alerts)} alerts")
        
        workbook.close()
        return receipts, self.processing_alerts
    
    def _validate_excel_structure(self, worksheet: Worksheet) -> None:
        """
        Validate Excel file has expected structure.
        
        Args:
            worksheet: Excel worksheet to validate
        """
        # Check header row exists
        if worksheet.max_row < 2:
            self.validation_errors.append("Excel file is empty or has no data rows")
            return
        
        # Get header row
        header_row = [cell.value for cell in worksheet[1]]
        
        # Required columns with Portuguese alternatives
        required_columns = {
            'contract': ['contract', 'contrato'],
            'name': ['name', 'nome'],
            'rent': ['rent', 'renda'],
            'rentdeposit': ['rentdeposit', 'deposit', 'caução', 'caucao'],
            'monthslate': ['monthslate', 'late', 'atraso'],
            'paidcurrentmonth': ['paidcurrentmonth', 'paid', 'mês caução', 'mes caucao']
        }
        
        header_lower = [str(h).lower().strip() if h else '' for h in header_row]
        
        missing_columns = []
        for req_col, alternatives in required_columns.items():
            # Check if any alternative appears in headers
            found = False
            for alt in alternatives:
                if any(alt in h for h in header_lower):
                    found = True
                    break
            if not found:
                missing_columns.append(req_col)
        
        if missing_columns:
            self.validation_errors.append(
                f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Check for at least some month columns (accept both text and numeric)
        month_count = sum(1 for h in header_row if h and str(h) in MONTH_COLUMNS)
        # Also check for numeric months (01-12)
        if month_count == 0:
            numeric_months = sum(1 for h in header_row if h and str(h).strip().isdigit() and 1 <= int(str(h).strip()) <= 12 and len(str(h).strip()) <= 2)
            if numeric_months == 0:
                self.validation_errors.append(
                    "No month columns found (expected Jan, Feb, Mar, etc. or 01, 02, 03, etc.)"
                )
            else:
                month_count = numeric_months
    
    def _parse_tenants(self, worksheet: Worksheet) -> List[TenantData]:
        """
        Parse tenant data from Excel worksheet.
        
        Args:
            worksheet: Excel worksheet to parse
        
        Returns:
            List of TenantData objects
        """
        tenants = []
        
        # Get header row to map column names
        header_row = [cell.value for cell in worksheet[1]]
        col_map = self._build_column_map(header_row)
        
        logger.info(f"Column mapping: {col_map}")
        logger.info(f"Total rows in worksheet: {worksheet.max_row}")
        
        # Parse data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row]
            
            # Skip landlord separator rows
            if self._is_landlord_separator_row(row_values):
                logger.debug(f"Row {row_idx}: Skipped (landlord separator)")
                continue
            
            # Skip empty rows
            if all(v is None or str(v).strip() == '' for v in row_values):
                logger.debug(f"Row {row_idx}: Skipped (empty)")
                continue
            
            # Parse tenant
            try:
                tenant = self._parse_tenant_row(row_values, col_map, row_idx)
                if tenant:
                    tenants.append(tenant)
                    logger.debug(f"Row {row_idx}: Parsed tenant '{tenant.name}'")
                else:
                    logger.debug(f"Row {row_idx}: Skipped (parse returned None)")
            except Exception as e:
                logger.warning(f"Row {row_idx}: Failed to parse - {str(e)}")
                self.validation_errors.append(f"Row {row_idx}: {str(e)}")
        
        return tenants
    
    def _build_column_map(self, header_row: List) -> dict:
        """
        Build mapping of column names to indices.
        
        Args:
            header_row: List of header cell values
        
        Returns:
            Dictionary mapping column purpose to index
        """
        col_map = {
            'contract': None,
            'name': None,
            'rent': None,
            'rent_deposit': None,
            'deposit_month_offset': None,
            'months_late': None,
            'paid_current_month': None,
            'month_columns': {}
        }
        
        for idx, header in enumerate(header_row):
            if not header:
                continue
            
            header_str = str(header).strip()
            header_lower = header_str.lower()
            
            # Normalize header for better matching (remove spaces, accents, etc.)
            header_normalized = header_lower.replace(' ', '').replace('ç', 'c').replace('ã', 'a').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
            
            # Map required columns - check most specific patterns first to avoid false matches
            # Priority order matters to avoid mismatches
            
            # 1. Contract column
            if col_map['contract'] is None and ('contrato' in header_normalized or 'contract' in header_normalized):
                col_map['contract'] = idx
                logger.debug(f"Mapped 'contract' to column {idx}: {header_str}")
            
            # 2. Name column  
            elif col_map['name'] is None and ('nome' in header_normalized or 'name' in header_normalized or 'tenant' in header_normalized):
                col_map['name'] = idx
                logger.debug(f"Mapped 'name' to column {idx}: {header_str}")
            
            # 3. Rent column (check it doesn't contain "deposit" to avoid rent_deposit column)
            elif col_map['rent'] is None and ('renda' in header_normalized or ('rent' in header_normalized and 'deposit' not in header_normalized)):
                col_map['rent'] = idx
                logger.debug(f"Mapped 'rent' to column {idx}: {header_str}")
            
            # 4. Mes Caucao column (most specific - check this before generic caucao)
            elif 'mescauca' in header_normalized or 'mescaucao' in header_normalized:
                # This is the Mes Caucao column - can be numeric offset OR Yes/No
                col_map['paid_current_month'] = idx
                col_map['deposit_month_offset'] = idx
                logger.debug(f"Mapped 'mes_caucao' (paid_current_month + deposit_month_offset) to column {idx}: {header_str}")
            
            # 5. Separate PaidCurrentMonth column (if exists)
            elif col_map['paid_current_month'] is None and 'paidcurrentmonth' in header_normalized:
                col_map['paid_current_month'] = idx
                logger.debug(f"Mapped 'paid_current_month' to column {idx}: {header_str}")
            
            # 6. Rent Deposit column
            elif col_map['rent_deposit'] is None and ('rentdeposit' in header_normalized or 'caucao' in header_normalized or 'caucao' in header_normalized or 'deposit' in header_normalized):
                col_map['rent_deposit'] = idx
                logger.debug(f"Mapped 'rent_deposit' to column {idx}: {header_str}")
            
            # 7. Months Late column
            elif col_map['months_late'] is None and ('atraso' in header_normalized or 'monthslate' in header_normalized or 'late' in header_normalized):
                col_map['months_late'] = idx
                logger.debug(f"Mapped 'months_late' to column {idx}: {header_str}")
            
            # 8. Month columns (text or numeric)
            # Check text month names first
            if header_str in MONTH_COLUMNS:
                month_num = MONTH_COLUMNS[header_str]
                col_map['month_columns'][month_num] = idx
                logger.debug(f"Mapped month {month_num} to column {idx}: {header_str}")
            # Also accept numeric months (01-12 or 1-12)
            elif header_str.strip().isdigit():
                month_num = int(header_str.strip())
                if 1 <= month_num <= 12:
                    col_map['month_columns'][month_num] = idx
                    logger.debug(f"Mapped month {month_num} to column {idx}: {header_str}")
        
        # Log final mapping
        logger.info(f"Final column mapping: contract={col_map['contract']}, name={col_map['name']}, rent={col_map['rent']}, "
                   f"rent_deposit={col_map['rent_deposit']}, months_late={col_map['months_late']}, "
                   f"month_columns={col_map['month_columns']}")
        
        return col_map
    
    def _is_landlord_separator_row(self, row_values: List) -> bool:
        """
        Check if row is a landlord separator.
        
        Args:
            row_values: List of cell values in row
        
        Returns:
            True if row is a landlord separator
        """
        # Check if first column contains text indicating landlord
        first_col = row_values[0] if row_values else None
        if not first_col:
            return False
        
        first_str = str(first_col).upper().strip()
        
        # Common patterns for landlord separators
        landlord_patterns = ['LANDLORD', 'PROPERTY OWNER', 'OWNER:', '[LANDLORD']
        
        return any(pattern in first_str for pattern in landlord_patterns)
    
    def _parse_tenant_row(
        self,
        row_values: List,
        col_map: dict,
        row_number: int
    ) -> Optional[TenantData]:
        """
        Parse a single tenant row.
        
        Args:
            row_values: List of cell values
            col_map: Column mapping dictionary
            row_number: Row number for error reporting
        
        Returns:
            TenantData object or None if row is invalid
        
        Raises:
            ValueError: If required data is missing or invalid
        """
        # Extract contract number
        contract_idx = col_map['contract']
        if contract_idx is None or contract_idx >= len(row_values):
            raise ValueError("Missing contract number column")
        
        contract = row_values[contract_idx]
        if not contract or str(contract).strip() == '':
            raise ValueError("Missing contract number")
        
        # Extract name
        name_idx = col_map['name']
        name = row_values[name_idx] if name_idx is not None and name_idx < len(row_values) else ''
        
        # Extract rent
        rent_idx = col_map['rent']
        if rent_idx is None or rent_idx >= len(row_values):
            raise ValueError("Missing rent column")
        
        rent = row_values[rent_idx]
        if not isinstance(rent, (int, float)) or rent <= 0:
            raise ValueError(f"Invalid rent amount: {rent}")
        
        # Extract rent_deposit
        rd_idx = col_map['rent_deposit']
        if rd_idx is None or rd_idx >= len(row_values):
            raise ValueError("Missing RentDeposit column")
        
        rent_deposit = row_values[rd_idx]
        if not isinstance(rent_deposit, int) or rent_deposit < 0:
            raise ValueError(f"Invalid RentDeposit: {rent_deposit} (must be non-negative integer)")
        
        # Extract deposit_month_offset (from "Mês Caução" column if it's numeric)
        # If Mes Caucao contains a number, use it; otherwise fall back to rent_deposit
        dmo_idx = col_map['deposit_month_offset']
        deposit_month_offset = rent_deposit  # Default fallback
        if dmo_idx is not None and dmo_idx < len(row_values):
            dmo_value = row_values[dmo_idx]
            # Only use it if it's a numeric value (not Yes/No string)
            if isinstance(dmo_value, (int, float)) and dmo_value >= 0:
                deposit_month_offset = int(dmo_value)
                logger.debug(f"Using deposit_month_offset from Mes Caucao: {deposit_month_offset}")
            else:
                logger.debug(f"Mes Caucao value '{dmo_value}' is not numeric, using rent_deposit: {rent_deposit}")
        else:
            logger.debug(f"No Mes Caucao column found, using rent_deposit: {rent_deposit}")
        
        # Extract months_late
        ml_idx = col_map['months_late']
        if ml_idx is None or ml_idx >= len(row_values):
            raise ValueError("Missing MonthsLate column")
        
        months_late = row_values[ml_idx]
        if not isinstance(months_late, int) or months_late < 0:
            raise ValueError(f"Invalid MonthsLate: {months_late} (must be non-negative integer)")
        
        # Extract paid_current_month
        # If Mes Caucao column contains numeric value, it's used for deposit offset
        # In that case, default paid_current_month to False
        pcm_idx = col_map['paid_current_month']
        paid_current_month = False  # Default
        if pcm_idx is not None and pcm_idx < len(row_values):
            paid_current_value = row_values[pcm_idx]
            # If it's not a number (i.e., it's Yes/No), parse it
            if not isinstance(paid_current_value, int):
                paid_current_month = self._parse_yes_no(paid_current_value)
        
        return TenantData(
            contract_number=str(contract).strip(),
            name=str(name).strip() if name else '',
            rent=float(rent),
            rent_deposit=int(rent_deposit),
            deposit_month_offset=int(deposit_month_offset),
            months_late=int(months_late),
            paid_current_month=paid_current_month,
            row_number=row_number
        )
    
    def _parse_yes_no(self, value) -> bool:
        """
        Parse Yes/No value to boolean.
        
        Args:
            value: Cell value
        
        Returns:
            True for Yes, False for No
        """
        if value is None:
            return False
        
        value_str = str(value).upper().strip()
        return value_str in ['YES', 'Y', 'TRUE', '1']
    
    def _prepare_receipt_records(
        self,
        tenants: List[TenantData],
        selected_month: int,
        selected_year: int,
        worksheet: Worksheet,
        col_map: dict
    ) -> List[ReceiptData]:
        """
        Prepare receipt data records from tenant information.
        Creates ReceiptData objects for CSV export - does NOT submit to portal.
        
        Args:
            tenants: List of tenant data
            selected_month: Target month (1-12)
            selected_year: Target year
            worksheet: Excel worksheet to read payment dates from
            col_map: Column mapping dictionary
        
        Returns:
            List of ReceiptData objects ready for CSV export
        """
        receipts = []
        
        # Month names for logging
        month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        # Get the column index for the selected month
        month_col_idx = col_map['month_columns'].get(selected_month)
        
        if month_col_idx is None:
            logger.error(f"Month {selected_month} column not found in Excel. Cannot process receipts without payment date column.")
            return receipts  # Return empty list
        
        for tenant in tenants:
            logger.debug(f"Processing contract {tenant.contract_number}: {tenant.name}")
            
            # Read the payment date from the month column cell for this tenant's row
            from datetime import date as date_cls
            import datetime as dt
            
            payment_date = None
            payment_day = None
            
            if month_col_idx is not None:
                # Get the cell value from the selected month column for this tenant's row
                # month_col_idx is 0-based, openpyxl uses 1-based columns
                try:
                    cell_value = worksheet.cell(row=tenant.row_number, column=month_col_idx + 1).value
                    logger.debug(f"  Reading from row {tenant.row_number}, column {month_col_idx + 1}, value: {cell_value}")
                    
                    # Skip this tenant if cell is empty - no payment date means no payment
                    if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
                        logger.info(f"  Skipping {tenant.name} (contract {tenant.contract_number}) - no payment date in {month_names[selected_month-1]} column")
                        continue
                    elif isinstance(cell_value, dt.datetime):
                        payment_day = cell_value.day
                        logger.debug(f"  Read payment date from cell: day {payment_day}")
                    elif isinstance(cell_value, dt.date):
                        payment_day = cell_value.day
                        logger.debug(f"  Read payment date from cell: day {payment_day}")
                    elif isinstance(cell_value, (int, float)):
                        # If it's a number, assume it's the day of month
                        payment_day = int(cell_value)
                        logger.debug(f"  Read payment day from cell: {payment_day}")
                    elif isinstance(cell_value, str):
                        # Try to parse as full date first (dd-mm-yyyy format)
                        try:
                            parsed_date = dt.datetime.strptime(cell_value.strip(), "%d-%m-%Y")
                            payment_day = parsed_date.day
                            logger.debug(f"  Parsed payment date from dd-mm-yyyy string: {cell_value} -> day {payment_day}")
                        except ValueError:
                            # Try to parse as number (day of month)
                            try:
                                payment_day = int(cell_value)
                                logger.debug(f"  Parsed payment day from string: {payment_day}")
                            except ValueError:
                                logger.warning(f"  Skipping {tenant.name} (contract {tenant.contract_number}) - could not parse payment date '{cell_value}'")
                                continue
                    else:
                        logger.warning(f"  Skipping {tenant.name} (contract {tenant.contract_number}) - unknown payment date cell type: {type(cell_value)}")
                        continue
                        
                except Exception as e:
                    logger.warning(f"  Skipping {tenant.name} (contract {tenant.contract_number}) - error reading payment date: {e}")
                    continue
            else:
                logger.warning(f"  Skipping {tenant.name} (contract {tenant.contract_number}) - no month column found for {month_names[selected_month-1]}")
                continue
            
            # Validate day is within valid range
            import calendar
            max_day = calendar.monthrange(selected_year, selected_month)[1]
            if payment_day < 1 or payment_day > max_day:
                logger.warning(f"  Skipping {tenant.name} (contract {tenant.contract_number}) - invalid payment day {payment_day} (must be 1-{max_day})")
                continue
            
            # Create payment date
            payment_date = date_cls(selected_year, selected_month, payment_day)
            
            # Calculate rent period based on deposit_month_offset (from "Mês Caução" column)
            # Formula: selected_month + mês_caução
            # Example: Feb(2) + 2 = April(4) - paying 2 months in advance
            from_month = selected_month + tenant.deposit_month_offset
            from_year = selected_year
            
            # Handle month overflow
            while from_month > 12:
                from_month -= 12
                from_year += 1
            
            from_date = date_cls(from_year, from_month, 1)
            
            # to_date is last day of from_month
            last_day = calendar.monthrange(from_year, from_month)[1]
            to_date = date_cls(from_year, from_month, last_day)
            
            # Create a basic receipt for this tenant
            receipt = ReceiptData(
                contract_id=tenant.contract_number,
                from_date=from_date,
                to_date=to_date,
                payment_date=payment_date,
                receipt_type="rent",
                value=tenant.rent,
                rent_deposit=tenant.rent_deposit,
                months_late=tenant.months_late
            )
            receipts.append(receipt)
            logger.debug(f"Created receipt record for {tenant.name} with payment date {payment_date}")
        
        logger.info(f"Prepared {len(receipts)} receipt records from {len(tenants)} tenants")
        return receipts
    
    def _calculate_expected_column(
        self,
        payment_date: date,
        rent_deposit: int,
        months_late: int,
        paid_current_month: bool
    ) -> str:
        """
        Calculate which column payment was expected in.
        
        Args:
            payment_date: Date payment was made
            rent_deposit: Months paid in advance
            months_late: Months behind
            paid_current_month: True if paying current month
        
        Returns:
            Expected month name (Jan, Feb, etc.)
        """
        if paid_current_month:
            expected_month = payment_date.month
        else:
            # Expected payment month = rent period + rent_deposit
            calculator = RentPeriodCalculator(
                payment_date, rent_deposit, months_late, paid_current_month
            )
            rent_from, _ = calculator.calculate()
            expected_month = rent_from.month
        
        # Convert month number to name
        month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        return month_names[expected_month] if 1 <= expected_month <= 12 else ''


def validate_excel_file(file_path: str) -> Tuple[bool, List[str]]:
    """
    Validate Excel file structure without full parsing.
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Tuple of (is_valid, error_list)
    """
    processor = LandlordExcelProcessor()
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        processor._validate_excel_structure(worksheet)
        workbook.close()
        
        return (len(processor.validation_errors) == 0, processor.validation_errors)
    
    except Exception as e:
        return (False, [f"Cannot open file: {str(e)}"])
