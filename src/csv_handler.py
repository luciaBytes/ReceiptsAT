"""
CSV Handler for processing receipt data from CSV files and Excel files.
Supports automatic conversion from Excel (.xlsx, .xls) to CSV format.
"""

import csv
import os
import tempfile
from datetime import datetime
from typing import List, Dict, Tuple
from dataclasses import dataclass

try:
    from .utils.logger import get_logger
except ImportError:
    # Fallback for when imported directly
    from utils.logger import get_logger

logger = get_logger(__name__)

@dataclass
class ReceiptData:
    """Data class for receipt information."""
    contract_id: str
    from_date: str
    to_date: str
    receipt_type: str
    value: float
    payment_date: str = ""  # Optional payment date field
    payment_date_defaulted: bool = False  # Track if payment_date was defaulted
    value_defaulted: bool = False  # Track if value was defaulted from contract
    receipt_type_defaulted: bool = False  # Track if receipt_type was defaulted to 'rent'
    row_number: int = 0

class CSVHandler:
    """Handles CSV file operations for receipt data."""
    
    REQUIRED_COLUMNS = ['contractId', 'fromDate', 'toDate']
    OPTIONAL_COLUMNS = ['value', 'paymentDate', 'receiptType']  # Value, paymentDate, and receiptType are optional
    
    # Column aliases for flexibility (alternative column names)
    COLUMN_ALIASES = {
        'contractId': ['contract_id', 'contract', 'contractid', 'contract_number'],
        'fromDate': ['from_date', 'start_date', 'startdate', 'from', 'start'],
        'toDate': ['to_date', 'end_date', 'enddate', 'to', 'end'],
        'receiptType': ['receipt_type', 'type', 'receipttype'],
        'paymentDate': ['payment_date', 'paid_date', 'paymentdate', 'paid', 'payment'],
        'value': ['amount', 'rent', 'price', 'total']
    }
    
    def __init__(self):
        self.receipts: List[ReceiptData] = []
        self.validation_errors: List[str] = []
        self.column_mapping: Dict[str, str] = {}  # Maps CSV columns to standard names
        self.temp_csv_file: str = None  # Track temporary CSV file from Excel conversion
    
    def _convert_excel_to_csv(self, excel_path: str) -> Tuple[bool, str, str]:
        """
        Convert Excel file to CSV format.
        
        Args:
            excel_path: Path to the Excel file (.xlsx or .xls)
            
        Returns:
            Tuple of (success, csv_path, error_message)
        """
        try:
            import openpyxl
            
            logger.info(f"Converting Excel file to CSV: {excel_path}")
            
            # Load the Excel workbook
            try:
                workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            except Exception as e:
                return False, "", f"Failed to open Excel file: {str(e)}"
            
            # Get the first sheet
            sheet = workbook.active
            logger.info(f"Reading from sheet: {sheet.title}")
            
            # Create a temporary CSV file
            temp_fd, temp_path = tempfile.mkstemp(suffix='.csv', text=True)
            os.close(temp_fd)  # Close the file descriptor
            
            try:
                # Write Excel data to CSV
                with open(temp_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                    csv_writer = csv.writer(csv_file)
                    
                    row_count = 0
                    for row in sheet.iter_rows(values_only=True):
                        # Skip completely empty rows
                        if all(cell is None or str(cell).strip() == '' for cell in row):
                            continue
                        
                        # Convert None values to empty strings
                        cleaned_row = ['' if cell is None else str(cell) for cell in row]
                        csv_writer.writerow(cleaned_row)
                        row_count += 1
                    
                    logger.info(f"Converted {row_count} rows from Excel to CSV")
                
                workbook.close()
                
                if row_count == 0:
                    os.unlink(temp_path)
                    return False, "", "Excel file is empty"
                
                self.temp_csv_file = temp_path
                return True, temp_path, ""
                
            except Exception as e:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                workbook.close()
                return False, "", f"Failed to write CSV: {str(e)}"
                
        except ImportError:
            return False, "", "openpyxl library not installed. Please install it with: pip install openpyxl"
        except Exception as e:
            return False, "", f"Excel conversion error: {str(e)}"
    
    def _cleanup_temp_csv(self):
        """Clean up temporary CSV file created from Excel conversion."""
        if self.temp_csv_file and os.path.exists(self.temp_csv_file):
            try:
                os.unlink(self.temp_csv_file)
                logger.info(f"Cleaned up temporary CSV file: {self.temp_csv_file}")
                self.temp_csv_file = None
            except Exception as e:
                logger.warning(f"Failed to clean up temporary CSV file: {str(e)}")
    
    def load_csv(self, file_path: str) -> Tuple[bool, List[str]]:
        """
        Load and validate CSV file or Excel file (with automatic conversion).
        Column order is flexible - uses header names.
        
        Args:
            file_path: Path to the CSV or Excel file (.csv, .xlsx, .xls)
            
        Returns:
            Tuple of (success, error_messages)
        """
        if not os.path.exists(file_path):
            return False, ["File does not exist"]
        
        # Clean up any previous temporary CSV file
        self._cleanup_temp_csv()
        
        # Check if file is Excel format
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.xlsx', '.xls']:
            logger.info(f"Detected Excel file: {file_path}")
            success, csv_path, error_msg = self._convert_excel_to_csv(file_path)
            if not success:
                return False, [error_msg]
            
            logger.info(f"Excel converted to temporary CSV: {csv_path}")
            # Use the converted CSV file for the rest of the processing
            file_path = csv_path
        
        try:
            self.receipts.clear()
            self.validation_errors.clear()
            self.column_mapping.clear()
            
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                # Try to detect dialect, with fallback to standard comma-separated
                sample = file.read(1024)
                file.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
                except csv.Error:
                    # Fallback to standard CSV dialect if detection fails
                    logger.warning("Could not detect CSV dialect, using default comma-separated format")
                    dialect = csv.excel  # Standard comma-separated
                
                reader = csv.DictReader(file, dialect=dialect)
                
                # Build column mapping from CSV headers to standard names
                success, mapping_errors = self._build_column_mapping(reader.fieldnames)
                if not success:
                    return False, mapping_errors
                
                logger.info(f"Column mapping: {self.column_mapping}")
                
                # Process each row
                for row_num, row in enumerate(reader, start=2):  # Start at 2 for header
                    try:
                        receipt = self._parse_row(row, row_num)
                        if receipt:
                            # Validate the receipt before adding it
                            temp_errors = []
                            original_errors = self.validation_errors.copy()
                            self.validation_errors = temp_errors
                            
                            self._validate_receipt(receipt)
                            
                            # If validation passed, add the receipt
                            if not temp_errors:
                                self.receipts.append(receipt)
                            
                            # Add any validation errors to the main list
                            original_errors.extend(temp_errors)
                            self.validation_errors = original_errors
                            
                    except Exception as e:
                        self.validation_errors.append(f"Row {row_num}: {str(e)}")
            
            # Final validation summary
            logger.info(f"CSV processing completed:")
            logger.info(f"  Total rows processed: {row_num - 1 if 'row_num' in locals() else 0}")
            logger.info(f"  Valid receipts loaded: {len(self.receipts)}")
            logger.info(f"  Validation errors: {len(self.validation_errors)}")
            
            if self.validation_errors:
                return False, self.validation_errors
            
            logger.info(f"Successfully loaded {len(self.receipts)} receipts from {file_path}")
            return True, []
            
        except Exception as e:
            logger.error(f"Error loading CSV file: {str(e)}")
            return False, [f"Error reading file: {str(e)}"]
    
    def _build_column_mapping(self, csv_columns: List[str]) -> Tuple[bool, List[str]]:
        """
        Build mapping from CSV column names to standard column names.
        Supports column aliases for flexibility.
        
        Args:
            csv_columns: List of column names from CSV header
            
        Returns:
            Tuple of (success, error_messages)
        """
        if not csv_columns:
            return False, ["CSV file has no columns"]
        
        # Normalize CSV column names (lowercase, strip spaces)
        normalized_csv_columns = [col.strip().lower() for col in csv_columns]
        
        # Find mapping for each required and optional column
        for standard_col in self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS:
            found = False
            
            # First try exact match (case insensitive)
            standard_col_lower = standard_col.lower()
            if standard_col_lower in normalized_csv_columns:
                original_col = csv_columns[normalized_csv_columns.index(standard_col_lower)]
                self.column_mapping[standard_col] = original_col
                found = True
            else:
                # Try aliases
                aliases = self.COLUMN_ALIASES.get(standard_col, [])
                for alias in aliases:
                    if alias.lower() in normalized_csv_columns:
                        original_col = csv_columns[normalized_csv_columns.index(alias.lower())]
                        self.column_mapping[standard_col] = original_col
                        found = True
                        break
            
            # Check if required column is missing
            if not found and standard_col in self.REQUIRED_COLUMNS:
                all_possible = [standard_col] + self.COLUMN_ALIASES.get(standard_col, [])
                return False, [f"Missing required column '{standard_col}'. Looked for: {', '.join(all_possible)}"]
        
        # Report found columns
        logger.info("Column mapping established:")
        for standard, csv_col in self.column_mapping.items():
            logger.info(f"  {standard} -> '{csv_col}'")
        
        return True, []
    
    def _normalize_date(self, date_str: str) -> str:
        """
        Normalize date string to YYYY-MM-DD format.
        Handles various formats including Excel datetime strings.
        
        Args:
            date_str: Date string in various formats
            
        Returns:
            Date in YYYY-MM-DD format
        """
        if not date_str:
            return ''
        
        date_str = str(date_str).strip()
        
        # If already in correct format, return as-is
        if len(date_str) == 10 and date_str.count('-') == 2:
            try:
                datetime.strptime(date_str, '%Y-%m-%d')
                return date_str
            except ValueError:
                pass
        
        # Try to parse common date formats (including with time)
        formats_to_try = [
            '%Y-%m-%d %H:%M:%S',      # Excel: 2025-11-01 00:00:00
            '%Y-%m-%d %H:%M:%S.%f',   # With microseconds
            '%d/%m/%Y',                # European: 01/11/2025
            '%m/%d/%Y',                # American: 11/01/2025
            '%Y/%m/%d',                # ISO variant: 2025/11/01
            '%d-%m-%Y',                # European with dashes
            '%Y-%m-%d',                # ISO standard
        ]
        
        for fmt in formats_to_try:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        # If all parsing fails, return original (will fail validation later)
        return date_str
    
    def _parse_row(self, row: Dict[str, str], row_num: int) -> ReceiptData:
        """Parse a single CSV row into ReceiptData using flexible column mapping."""
        try:
            # Use column mapping to get values
            def get_mapped_value(standard_col: str) -> str:
                csv_col = self.column_mapping.get(standard_col)
                if csv_col and csv_col in row:
                    return row[csv_col].strip()
                return ''
            
            # Handle optional value with fallback
            value_str = get_mapped_value('value')
            value_defaulted = False
            if value_str:
                # Parse provided value
                value = float(value_str.replace(',', '.'))
            else:
                # Use fallback value - will be filled from contract data later
                # Use -1.0 to indicate missing value (easier to detect than 0.0)
                value = -1.0
                value_defaulted = True
            
            # Handle and normalize dates (strip time if present)
            from_date = self._normalize_date(get_mapped_value('fromDate'))
            to_date = self._normalize_date(get_mapped_value('toDate'))
            payment_date = self._normalize_date(get_mapped_value('paymentDate'))
            
            payment_date_defaulted = False
            if not payment_date:
                # Payment date is required - raise error
                raise ValueError(f"Row {row_num}: Payment date is required but missing")
            
            # Handle receipt type (optional - defaults to 'rent' if not provided)
            receipt_type = get_mapped_value('receiptType')
            receipt_type_defaulted = False
            if not receipt_type:
                # Default to 'rent' if not provided
                receipt_type = 'rent'
                receipt_type_defaulted = True
            
            # Validate date formats (should now all be YYYY-MM-DD)
            for date_field, date_val in [('fromDate', from_date), ('toDate', to_date), ('paymentDate', payment_date)]:
                if date_val:
                    try:
                        datetime.strptime(date_val, '%Y-%m-%d')
                    except ValueError:
                        raise ValueError(f"Row {row_num}: {date_field} has invalid format '{date_val}', expected YYYY-MM-DD")
            
            return ReceiptData(
                contract_id=get_mapped_value('contractId'),
                from_date=from_date,
                to_date=to_date,
                receipt_type=receipt_type,
                value=value,
                payment_date=payment_date,
                payment_date_defaulted=payment_date_defaulted,
                value_defaulted=value_defaulted,
                receipt_type_defaulted=receipt_type_defaulted,
                row_number=row_num
            )
        except ValueError as e:
            raise ValueError(f"Invalid data format: {str(e)}")
    
    def _validate_data(self):
        """Validate all receipt data."""
        for receipt in self.receipts:
            self._validate_receipt(receipt)
    
    def _validate_receipt(self, receipt: ReceiptData):
        """Validate a single receipt."""
        # Validate contract ID
        if not receipt.contract_id:
            self.validation_errors.append(f"Row {receipt.row_number}: Contract ID cannot be empty")
        
        # Validate dates
        try:
            from_date = datetime.strptime(receipt.from_date, '%Y-%m-%d')
            to_date = datetime.strptime(receipt.to_date, '%Y-%m-%d')
            payment_date = datetime.strptime(receipt.payment_date, '%Y-%m-%d')
            current_date = datetime.now()
            
            # Check if from_date is later than to_date
            if from_date > to_date:
                self.validation_errors.append(
                    f"Row {receipt.row_number}: From date ({receipt.from_date}) cannot be later than to date ({receipt.to_date})"
                )
            
            # Check if payment date is in the future
            if payment_date.date() > current_date.date():
                self.validation_errors.append(
                    f"Row {receipt.row_number}: Payment date ({receipt.payment_date}) cannot be in the future"
                )
                
        except ValueError as e:
            self.validation_errors.append(
                f"Row {receipt.row_number}: Invalid date format. Use YYYY-MM-DD. Error: {str(e)}"
            )
        
        # Validate value (allow -1.0 as fallback indicator, but other negative values are invalid)
        if receipt.value < 0 and receipt.value != -1.0:
            self.validation_errors.append(
                f"Row {receipt.row_number}: Value cannot be negative (got {receipt.value})"
            )
        elif receipt.value == 0.0 and not receipt.value_defaulted:
            # Warn about explicit zero values (might be intentional but unusual)
            logger.warning(f"Row {receipt.row_number}: Value is 0.0 - verify this is intentional")
        
        # Validate receipt type
        if not receipt.receipt_type:
            self.validation_errors.append(
                f"Row {receipt.row_number}: Receipt type cannot be empty"
            )
    
    def get_receipts(self) -> List[ReceiptData]:
        """Get the list of loaded receipts."""
        return self.receipts.copy()
    
    def clear_data(self):
        """Clear all loaded CSV data."""
        self.receipts = []
        self.validation_errors = []
        self.column_mapping = {}
        if self.temp_csv_file and os.path.exists(self.temp_csv_file):
            try:
                os.remove(self.temp_csv_file)
            except:
                pass
        self.temp_csv_file = None
    
    def filter_receipts_by_contracts(self, valid_contract_ids: List[str]) -> int:
        """
        Filter receipts to only include those with valid contract IDs.
        This is used after validation to ensure we only process receipts
        for contracts that exist on the platform.
        
        Args:
            valid_contract_ids: List of contract IDs that were validated on the platform
            
        Returns:
            Number of receipts removed
        """
        if not valid_contract_ids:
            logger.warning("No valid contract IDs provided - all receipts will be removed")
            removed_count = len(self.receipts)
            self.receipts.clear()
            return removed_count
        
        # Convert to strings for comparison
        valid_ids_set = set(str(cid).strip() for cid in valid_contract_ids)
        
        # Filter receipts
        original_count = len(self.receipts)
        self.receipts = [
            receipt for receipt in self.receipts
            if str(receipt.contract_id).strip() in valid_ids_set
        ]
        removed_count = original_count - len(self.receipts)
        
        logger.info(f"Filtered receipts: kept {len(self.receipts)} receipts, removed {removed_count} receipts")
        
        if removed_count > 0:
            logger.info(f"Removed receipts for contracts not found on platform: {removed_count}")
        
        return removed_count
    
    def get_contract_ids(self) -> List[str]:
        """
        Extract unique contract IDs from loaded receipt data.
        
        Returns:
            List of unique contract IDs
        """
        if not self.receipts:
            logger.warning("No receipt data loaded")
            return []
        
        contract_ids = list(set(receipt.contract_id for receipt in self.receipts))
        logger.info(f"Extracted {len(contract_ids)} unique contract IDs from CSV: {contract_ids}")
        return sorted(contract_ids)
    
    def get_contracts_summary(self) -> Dict[str, int]:
        """
        Get summary of contract IDs and their receipt counts.
        
        Returns:
            Dictionary mapping contract IDs to number of receipts
        """
        summary = {}
        for receipt in self.receipts:
            contract_id = receipt.contract_id
            summary[contract_id] = summary.get(contract_id, 0) + 1
        
        logger.info(f"Contract summary: {summary}")
        return summary

    def export_report(self, report_data: List[Dict], file_path: str) -> bool:
        """
        Export report data to CSV.
        
        Args:
            report_data: List of dictionaries containing report data
            file_path: Path to save the report
            
        Returns:
            Success status
        """
        try:
            if not report_data:
                return False
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as file:
                writer = csv.DictWriter(file, fieldnames=report_data[0].keys())
                writer.writeheader()
                writer.writerows(report_data)
            
            logger.info(f"Report exported to {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting report: {str(e)}")
            return False
    
    def export_session_report(self, report_data: List[Dict], file_path: str, append: bool = True) -> bool:
        """
        Export session report data to CSV with option to append to existing file.
        
        This method is designed for session-based exports where multiple processing runs
        should accumulate in the same file. The header is only written for new files.
        
        Args:
            report_data: List of dictionaries containing report data
            file_path: Path to save the report
            append: If True and file exists, append data without header; 
                   If False, overwrite file with header
            
        Returns:
            Success status
        """
        try:
            if not report_data:
                logger.warning("No report data to export")
                return False
            
            # Check if file exists
            file_exists = os.path.exists(file_path)
            
            # Determine mode: append or write
            if append and file_exists:
                # Append mode: add data without header
                mode = 'a'
                write_header = False
                logger.info(f"Appending {len(report_data)} rows to existing file: {file_path}")
            else:
                # Write mode: create new file with header
                mode = 'w'
                write_header = True
                logger.info(f"Creating new export file with {len(report_data)} rows: {file_path}")
            
            with open(file_path, mode, newline='', encoding='utf-8-sig') as file:
                writer = csv.DictWriter(file, fieldnames=report_data[0].keys())
                
                # Write header only for new files
                if write_header:
                    writer.writeheader()
                
                # Write all data rows
                writer.writerows(report_data)
            
            logger.info(f"Session report exported successfully to {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting session report: {str(e)}")
            return False
    
    def export_session_report_excel(self, report_data: List[Dict], file_path: str, append: bool = True) -> bool:
        """
        Export session report data to Excel with option to append to existing file.
        
        Args:
            report_data: List of dictionaries containing report data
            file_path: Path to save the report (.xlsx)
            append: If True and file exists, append data to existing sheet;
                   If False, overwrite file
            
        Returns:
            Success status
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            if not report_data:
                logger.warning("No report data to export")
                return False
            
            # Check if file exists
            file_exists = os.path.exists(file_path)
            
            if append and file_exists:
                # Load existing workbook and append data
                logger.info(f"Appending {len(report_data)} rows to existing Excel file: {file_path}")
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                
                # Find the next empty row
                next_row = sheet.max_row + 1
                
                # Append data rows
                for data_row in report_data:
                    for col_idx, key in enumerate(data_row.keys(), start=1):
                        cell = sheet.cell(row=next_row, column=col_idx)
                        cell.value = data_row[key]
                    next_row += 1
            else:
                # Create new workbook
                logger.info(f"Creating new Excel file with {len(report_data)} rows: {file_path}")
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Session Report"
                
                # Write header
                headers = list(report_data[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data rows
                for row_idx, data_row in enumerate(report_data, start=2):
                    for col_idx, key in enumerate(headers, start=1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.value = data_row[key]
                
                # Auto-size columns
                for col_idx in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = len(str(headers[col_idx - 1]))
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(file_path)
            logger.info(f"Session report exported successfully to {file_path}")
            return True
            
        except ImportError:
            logger.error("openpyxl library not installed. Cannot export to Excel.")
            return False
        except Exception as e:
            logger.error(f"Error exporting session report to Excel: {str(e)}")
            return False
    
    def export_errors_report(self, results: List, file_path: str) -> bool:
        """
        Export failed receipts with detailed error information.
        
        Args:
            results: List of ProcessingResult objects
            file_path: Path to save the error report
            
        Returns:
            Success status
        """
        try:
            # Filter for failed receipts only
            failed_results = [r for r in results if not r.success and r.status == "Failed"]
            
            if not failed_results:
                logger.warning("No failed receipts to export")
                return False
            
            # Prepare error report data
            error_data = []
            for result in failed_results:
                error_row = {
                    'Contract Number': result.contract_id,
                    'Tenant Name': result.tenant_name or 'Unknown',
                    'Value': f"€{result.value:.2f}" if result.value else '',
                    'Period From': result.from_date,
                    'Period To': result.to_date,
                    'Payment Date': result.payment_date,
                    'Error Message': result.error_message,
                    'Field Errors': result.field_errors,
                    'Timestamp': result.timestamp
                }
                error_data.append(error_row)
            
            # Write to CSV
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as file:
                fieldnames = ['Contract Number', 'Tenant Name', 'Value', 'Period From', 'Period To', 
                             'Payment Date', 'Error Message', 'Field Errors', 'Timestamp']
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(error_data)
            
            logger.info(f"Error report exported to {file_path} with {len(error_data)} failed receipts")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting errors report: {str(e)}")
            return False
    
    def export_errors_report_excel(self, results: List, file_path: str) -> bool:
        """
        Export failed receipts with detailed error information to Excel.
        
        Args:
            results: List of ProcessingResult objects
            file_path: Path to save the error report (.xlsx)
            
        Returns:
            Success status
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Filter for failed receipts only
            failed_results = [r for r in results if not r.success and r.status == "Failed"]
            
            if not failed_results:
                logger.warning("No failed receipts to export")
                return False
            
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Error Report"
            
            # Define headers
            headers = ['Contract Number', 'Tenant Name', 'Value', 'Period From', 'Period To', 
                      'Payment Date', 'Error Message', 'Field Errors', 'Timestamp']
            
            # Write header row
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, result in enumerate(failed_results, start=2):
                sheet.cell(row=row_idx, column=1).value = result.contract_id
                sheet.cell(row=row_idx, column=2).value = result.tenant_name or 'Unknown'
                sheet.cell(row=row_idx, column=3).value = f"€{result.value:.2f}" if result.value else ''
                sheet.cell(row=row_idx, column=4).value = result.from_date
                sheet.cell(row=row_idx, column=5).value = result.to_date
                sheet.cell(row=row_idx, column=6).value = result.payment_date
                sheet.cell(row=row_idx, column=7).value = result.error_message
                sheet.cell(row=row_idx, column=8).value = result.field_errors
                sheet.cell(row=row_idx, column=9).value = result.timestamp
            
            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(headers[col_idx - 1]))
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 60)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(file_path)
            logger.info(f"Error report exported to {file_path} with {len(failed_results)} failed receipts")
            return True
            
        except ImportError:
            logger.error("openpyxl library not installed. Cannot export to Excel.")
            return False
        except Exception as e:
            logger.error(f"Error exporting errors report to Excel: {str(e)}")
            return False
