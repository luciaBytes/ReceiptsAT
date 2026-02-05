"""
Smart Import tab - Full-featured workflow with Excel and CSV processing.

This tab provides the complete workflow including:
- Excel pre-processing for monthly receipt generation (optional)
- CSV file import (standard workflow)
- Receipt validation and processing
- Progress tracking and logging
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from typing import Callable, List, Optional, Dict, Any
from datetime import datetime
import threading
import os
import sys
import zipfile
import xml.etree.ElementTree as ET
import openpyxl

try:
    from excel_preprocessor import LandlordExcelProcessor, ProcessingAlert, ReceiptData
    from csv_handler import CSVHandler
    from web_client import WebClient
    from receipt_processor import ReceiptProcessor
    from utils.logger import get_logger
    from utils.multilingual_localization import get_text
    from gui.themed_button import ThemedButton
except ImportError:
    from src.excel_preprocessor import LandlordExcelProcessor, ProcessingAlert, ReceiptData
    from src.csv_handler import CSVHandler
    from src.web_client import WebClient
    from src.receipt_processor import ReceiptProcessor
    from src.utils.logger import get_logger
    from src.utils.multilingual_localization import get_text
    from src.gui.themed_button import ThemedButton

logger = get_logger(__name__)


class SmartImportTab(tk.Frame):
    """Smart Import tab with full Excel and CSV processing capabilities."""
    
    def __init__(self, parent, csv_handler: CSVHandler, web_client: WebClient, 
                 processor: ReceiptProcessor, mode_var: tk.StringVar, 
                 dry_run_var: tk.BooleanVar, on_log: Callable):
        """
        Initialize Smart Import tab.
        
        Args:
            parent: Parent widget (notebook)
            csv_handler: CSV handler instance
            web_client: Web client instance
            processor: Receipt processor instance
            mode_var: Processing mode variable (bulk/step)
            dry_run_var: Dry run flag variable
            on_log: Logging callback function
        """
        super().__init__(parent, bg='#1e293b')
        
        # Store references
        self.csv_handler = csv_handler
        self.web_client = web_client
        self.processor = processor
        self.mode_var = mode_var
        self.dry_run_var = dry_run_var
        self.on_log = on_log
        
        # Initialize Excel processor
        self.excel_processor = LandlordExcelProcessor()
        
        # Variables
        self.excel_file_path = tk.StringVar()
        self.csv_file_path = tk.StringVar()
        self.selected_month = tk.IntVar(value=datetime.now().month)
        self.selected_sheet = tk.StringVar()  # Sheet name represents the year
        self.available_sheets = []  # List of sheet names from Excel
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value=get_text('STATUS_READY'))
        
        # State
        self.current_receipts: List[ReceiptData] = []
        self.current_alerts: List[ProcessingAlert] = []
        self.processing_thread = None
        self.stop_requested = False
        self.cached_validation_report: Optional[Dict[str, Any]] = None  # Cache validation results
        
        self._setup_gui()
    
    def _setup_gui(self):
        """Setup the tab GUI components."""
        # Create canvas and scrollbar for scrolling
        canvas = tk.Canvas(self, bg='#1e293b', highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#1e293b')
        
        # Configure canvas scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True, padx=12, pady=12)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel for scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Configure grid for scrollable frame instead of self
        scrollable_frame.columnconfigure(0, weight=1)
        scrollable_frame.rowconfigure(2, weight=1)  # Processing section expands for log
        
        # Excel Import Section with dark styling
        excel_frame = tk.LabelFrame(scrollable_frame, text="  📁 Excel Import  ", 
                                   bg='#1e293b', fg='#3b82f6',
                                   font=('Segoe UI', 9), relief='solid', borderwidth=1,
                                   padx=12, pady=8)
        excel_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        excel_frame.columnconfigure(1, weight=1)  # Excel file entry expands
        
        # Excel file selection
        tk.Label(excel_frame, text="Excel File:", bg='#1e293b', fg='#ffffff', font=('Segoe UI', 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        ttk.Entry(excel_frame, textvariable=self.excel_file_path, state="readonly", font=('Segoe UI', 9), foreground='black').grid(
            row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 6)
        )
        
        ThemedButton(excel_frame, style='secondary', text="📁 Browse...", 
                    command=self._browse_excel).grid(row=0, column=2, padx=(0, 6))
        
        # Data Selection Section with dark styling
        data_frame = tk.LabelFrame(scrollable_frame, text="  📅 Data Selection  ", 
                                   bg='#1e293b', fg='#3b82f6',
                                   font=('Segoe UI', 9), relief='solid', borderwidth=1,
                                   padx=12, pady=8)
        data_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Single row with month, year, and download button
        selection_row = tk.Frame(data_frame, bg='#1e293b')
        selection_row.pack(fill=tk.X)
        
        tk.Label(selection_row, text="Month:", bg='#1e293b', fg='#ffffff', font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        months = ["January", "February", "March", "April", "May", "June", 
                 "July", "August", "September", "October", "November", "December"]
        self.month_combo = ttk.Combobox(selection_row, values=months, width=12, state="readonly", foreground='black')
        self.month_combo.current(self.selected_month.get() - 1)
        self.month_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.month_combo.bind("<<ComboboxSelected>>", self._on_month_year_selected)
        
        tk.Label(selection_row, text="Year Tab:", bg='#1e293b', fg='#ffffff', font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 5))
        self.sheet_combo = ttk.Combobox(selection_row, textvariable=self.selected_sheet, width=10, state="readonly", foreground='black')
        self.sheet_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_month_year_selected)
        
        self.generate_csv_button = ThemedButton(selection_row, style='primary', text="📊 Generate CSV Data", 
                                               command=self._process_excel, state="disabled")
        self.generate_csv_button.pack(side=tk.LEFT, padx=(15, 0))
        
        self.download_csv_button = ThemedButton(selection_row, style='secondary', text="⬇ Download CSV", 
                                               command=self._download_csv, state="disabled")
        self.download_csv_button.pack(side=tk.LEFT, padx=(5, 0))
        
        # CSV File section (from Quick Import)
        csv_frame = tk.LabelFrame(scrollable_frame, text="  📄 " + get_text('CSV_FILE_SECTION') + "  ",
                                 bg='#1e293b', fg='#3b82f6',
                                 font=('Segoe UI', 9), relief='solid', borderwidth=1,
                                 padx=12, pady=8)
        csv_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # File selection row
        file_row = tk.Frame(csv_frame, bg='#1e293b')
        file_row.pack(fill=tk.X)
        
        tk.Label(file_row, text=get_text('FILE_LABEL'),
                bg='#1e293b', fg='#ffffff',
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 8))
        
        file_entry = ttk.Entry(file_row, textvariable=self.csv_file_path, state="readonly", foreground='black')
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        
        ThemedButton(file_row, style='secondary', text="📁 " + get_text('BROWSE_BUTTON'), 
                    command=self._browse_csv_file).pack(side=tk.LEFT, padx=(0, 6))
        
        # Processing Section with dark styling - buttons and status only
        processing_frame = tk.LabelFrame(scrollable_frame, text="  ⏱ Processing  ", 
                                        bg='#1e293b', fg='#3b82f6',
                                        font=('Segoe UI', 9), relief='solid', borderwidth=1,
                                        padx=12, pady=8)
        processing_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        processing_frame.columnconfigure(0, weight=1)
        
        # Processing control buttons
        button_frame = tk.Frame(processing_frame, bg='#1e293b')
        button_frame.grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        
        # Validate button (Tertiary style)
        self.validate_button = ThemedButton(button_frame, style='tertiary',
                                           text="✓ " + get_text('VALIDATE_CONTRACTS_BUTTON'), 
                                           command=self._validate_contracts, state="disabled")
        self.validate_button.pack(side=tk.LEFT, padx=(0, 6))
        
        # Review CSV button
        self.review_button = ThemedButton(button_frame, style='secondary', text="🔍 Review CSV", 
                                         command=self._review_csv, state="disabled")
        self.review_button.pack(side=tk.LEFT, padx=(0, 6))
        
        # Process receipts button (Primary style - main action)
        self.start_button = ThemedButton(button_frame, style='primary',
                                        text="▶ " + get_text('PROCESS_RECEIPTS_BUTTON'), 
                                        command=self._start_processing, state="disabled")
        self.start_button.pack(side=tk.LEFT, padx=(0, 6))
        
        # Stop button (Danger style)
        self.stop_button = ThemedButton(button_frame, style='danger',
                                       text="⏹ " + get_text('CANCEL_BUTTON'),
                                       command=self._stop_processing, state="disabled")
        self.stop_button.pack(side=tk.LEFT, padx=(0, 6))
        
        # Status field (not subsection)
        status_container = tk.Frame(processing_frame, bg='#1e293b')
        status_container.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 0))
        status_container.columnconfigure(0, weight=1)
        
        tk.Label(status_container, text="Status:", bg='#1e293b', fg='#94a3b8', 
                font=('Segoe UI', 9, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0, 4))
        
        self.progress_bar = ttk.Progressbar(status_container, variable=self.progress_var, 
                                           maximum=100, mode='determinate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 4))
        
        status_label = tk.Label(status_container, textvariable=self.status_var,
                               bg='#1e293b', fg='#ffffff',
                               font=('Segoe UI', 9))
        status_label.grid(row=2, column=0, sticky=tk.W)
        
        # Processing Results section (from Quick Import) with log
        results_frame = tk.LabelFrame(scrollable_frame, text="  ✓ Results  ",
                                     bg='#1e293b', fg='#3b82f6',
                                     font=('Segoe UI', 9), relief='solid', borderwidth=1,
                                     padx=12, pady=8)
        results_frame.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(1, weight=1)  # Log expands
        
        # Export buttons (before logs) - aligned to the left
        export_buttons_frame = tk.Frame(results_frame, bg='#1e293b')
        export_buttons_frame.grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        
        self.export_results_button = ThemedButton(export_buttons_frame, style='secondary',
                                                  text="💾 Export Results", 
                                                  command=self._export_all_results)
        self.export_results_button.pack(side=tk.LEFT, padx=(0, 6))
        
        self.export_logs_button = ThemedButton(export_buttons_frame, style='tertiary',
                                                  text="📊 Export Logs", 
                                                  command=self._export_logs)
        self.export_logs_button.pack(side=tk.LEFT, padx=(0, 6))
        
        self.verify_export_button = ThemedButton(export_buttons_frame, style='success',
                                                 text="✅ Verify Receipts", 
                                                 command=self._verify_and_export_receipts)
        self.verify_export_button.pack(side=tk.LEFT)
        
        self.results_label = tk.Label(results_frame, text="No receipts processed yet",
                                           bg='#1e293b', fg='#94a3b8',
                                           font=('Segoe UI', 9))
        self.results_label.grid(row=1, column=0, sticky=tk.W, pady=(0, 8))
        
        # Log field in results section
        log_container = tk.Frame(results_frame, bg='#1e293b')
        log_container.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_container.columnconfigure(0, weight=1)
        log_container.rowconfigure(1, weight=1)
        
        tk.Label(log_container, text="Log:", bg='#1e293b', fg='#94a3b8', 
                font=('Segoe UI', 9, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0, 4))
        
        self.log_text = scrolledtext.ScrolledText(log_container, height=10, width=80, wrap=tk.WORD,
                                                  font=('Consolas', 9),
                                                  bg='#0f172a',
                                                  fg='#94a3b8',
                                                  relief='sunken',
                                                  borderwidth=2,
                                                  highlightthickness=0,
                                                  padx=10,
                                                  pady=10)
        self.log_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Setup logging handler for Smart Import
        self._setup_logging_handler()
    
    def _setup_logging_handler(self):
        """Setup logging handler to capture logs in the log text widget."""
        import logging
        
        class GUILogHandler(logging.Handler):
            """Custom logging handler that writes to a tkinter Text widget."""
            def __init__(self, text_widget, root_widget):
                super().__init__()
                self.text_widget = text_widget
                self.root_widget = root_widget
            
            def emit(self, record):
                """Emit a log record to the text widget."""
                try:
                    msg = self.format(record)
                    # Schedule the GUI update on the main thread
                    self.root_widget.after(0, lambda: self._append_log(msg))
                except Exception:
                    pass
            
            def _append_log(self, msg):
                """Append log message to text widget (must run on main thread)."""
                try:
                    self.text_widget.insert(tk.END, msg + '\n')
                    self.text_widget.see(tk.END)
                    # Limit log size to prevent memory issues
                    lines = int(self.text_widget.index('end-1c').split('.')[0])
                    if lines > 1000:
                        self.text_widget.delete('1.0', f'{lines-1000}.0')
                except Exception:
                    pass
        
        # Create and configure the handler
        gui_handler = GUILogHandler(self.log_text, self.winfo_toplevel())
        gui_handler.setLevel(logging.DEBUG)
        
        # Use a detailed formatter
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%H:%M:%S'
        )
        gui_handler.setFormatter(formatter)
        
        # Add handler to receipts_app logger
        receipts_logger = logging.getLogger('receipts_app')
        if gui_handler not in receipts_logger.handlers:
            receipts_logger.addHandler(gui_handler)
            self._log_handler = gui_handler  # Store reference for cleanup
    
    def _create_compatible_excel(self, source_path, dest_path, sheet_name):
        """Create an openpyxl-compatible Excel file from a problematic source file."""
        try:
            # Create new workbook
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # Remove default sheet
            new_ws = new_wb.create_sheet(sheet_name)
            
            # Read all data from source ZIP
            with zipfile.ZipFile(source_path, 'r') as zip_ref:
                # Load shared strings
                shared_strings = []
                if 'xl/sharedStrings.xml' in zip_ref.namelist():
                    with zip_ref.open('xl/sharedStrings.xml') as ss_file:
                        ss_tree = ET.parse(ss_file)
                        for elem in ss_tree.iter():
                            if elem.tag.endswith('t'):
                                if elem.text:
                                    shared_strings.append(elem.text)
                
                # Read worksheet data
                sheet_xml_path = 'xl/worksheets/sheet1.xml'
                if sheet_xml_path in zip_ref.namelist():
                    with zip_ref.open(sheet_xml_path) as sheet_file:
                        sheet_tree = ET.parse(sheet_file)
                        sheet_root = sheet_tree.getroot()
                        
                        # Process each row
                        row_num = 1
                        for row_elem in sheet_root.iter():
                            if not row_elem.tag.endswith('row'):
                                continue
                            
                            col_num = 1
                            for cell_elem in row_elem:
                                if not cell_elem.tag.endswith('c'):
                                    continue
                                
                                # Get cell value
                                cell_type = cell_elem.attrib.get('t', '')
                                v_elem = None
                                for child in cell_elem:
                                    if child.tag.endswith('v'):
                                        v_elem = child
                                        break
                                
                                if v_elem is not None and v_elem.text:
                                    if cell_type == 's':
                                        # Shared string
                                        idx = int(v_elem.text)
                                        if 0 <= idx < len(shared_strings):
                                            value = shared_strings[idx]
                                        else:
                                            value = v_elem.text
                                    else:
                                        # Try to convert to number
                                        try:
                                            value = float(v_elem.text)
                                            if value.is_integer():
                                                value = int(value)
                                        except:
                                            value = v_elem.text
                                    
                                    # Write to new sheet
                                    new_ws.cell(row=row_num, column=col_num, value=value)
                                
                                col_num += 1
                            row_num += 1
            
            # Save the new workbook
            new_wb.save(dest_path)
            self.on_log("INFO", f"Created compatible Excel file with {row_num-1} rows")
            
        except Exception as e:
            self.on_log("ERROR", f"Failed to create compatible file: {e}")
            import traceback
            self.on_log("ERROR", traceback.format_exc())
            raise
    
    def _on_month_year_selected(self, event=None):
        """Handle month/year selection - update selection only, don't auto-process."""
        self.selected_month.set(self.month_combo.current() + 1)
        # Don't auto-process - wait for user to click Process button
    
    def _browse_excel(self):
        """Browse for Excel file, detect month columns and year from structure."""
        file_path = filedialog.askopenfilename(
            title="Select Landlord Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            self.excel_file_path.set(file_path)
            
            # Load and analyze file structure
            try:
                file_ext = os.path.splitext(file_path)[1].lower()
                self.on_log("INFO", f"File extension: {file_ext}")
                
                # First, inspect the xlsx file structure directly using zipfile
                try:
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        file_list = zip_ref.namelist()
                        self.on_log("INFO", f"ZIP contents ({len(file_list)} files): {file_list[:10]}...")
                        
                        # Try to read workbook.xml to get sheet names
                        if 'xl/workbook.xml' in file_list:
                            with zip_ref.open('xl/workbook.xml') as xml_file:
                                content = xml_file.read()
                                self.on_log("DEBUG", f"workbook.xml size: {len(content)} bytes")
                                
                                tree = ET.fromstring(content)
                                
                                # Try multiple namespace approaches
                                # Method 1: With namespace
                                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                sheets = tree.findall('.//main:sheet', ns)
                                self.on_log("DEBUG", f"Method 1 (with namespace): found {len(sheets)} sheets")
                                
                                # Method 2: Without namespace (any sheet tag)
                                if not sheets:
                                    # Find all elements ending with 'sheet'
                                    for elem in tree.iter():
                                        if elem.tag.endswith('sheet'):
                                            sheets.append(elem)
                                    self.on_log("DEBUG", f"Method 2 (any 'sheet' tag): found {len(sheets)} sheets")
                                
                                # Method 3: Direct search for any 'sheets' parent
                                if not sheets:
                                    for elem in tree.iter():
                                        if elem.tag.endswith('sheets'):
                                            sheets = list(elem)
                                            self.on_log("DEBUG", f"Method 3 (children of 'sheets'): found {len(sheets)} sheets")
                                            break
                                
                                self.available_sheets = [sheet.attrib.get('name', f"Sheet{i+1}") 
                                                       for i, sheet in enumerate(sheets) 
                                                       if hasattr(sheet, 'attrib')]
                                self.on_log("INFO", f"✓ Extracted {len(self.available_sheets)} sheets from XML: {self.available_sheets}")
                        else:
                            self.on_log("ERROR", "xl/workbook.xml not found in file!")
                            
                except Exception as zip_error:
                    self.on_log("ERROR", f"ZIP inspection failed: {zip_error}")
                    import traceback
                    self.on_log("ERROR", traceback.format_exc())
                    self.available_sheets = []
                
                # If ZIP method worked, continue; otherwise try openpyxl as fallback
                if not self.available_sheets:
                    self.on_log("WARNING", "Trying openpyxl as fallback...")
                    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False, keep_links=False)
                    self.available_sheets = wb.sheetnames
                    wb.close()
                    self.on_log("INFO", f"Openpyxl found {len(self.available_sheets)} sheets: {self.available_sheets}")
                
                if not self.available_sheets:
                    messagebox.showerror("Invalid File", 
                                       "Excel file has no sheets or is corrupted.\n\n" +
                                       f"File: {os.path.basename(file_path)}\n" +
                                       "Please verify the file opens correctly in Excel.")
                    self.excel_file_path.set("")
                    return
                
                # Detect year from sheet names (sheet names are exactly 4-digit years like 2026, 2027)
                year_sheets = []
                self.on_log("INFO", f"Scanning {len(self.available_sheets)} sheets for year detection...")
                
                for sheet_name in self.available_sheets:
                    sheet_str = str(sheet_name).strip()
                    self.on_log("DEBUG", f"  Sheet: {repr(sheet_name)} -> Type: {type(sheet_name).__name__} -> String: '{sheet_str}' -> isdigit: {sheet_str.isdigit()} -> len: {len(sheet_str)}")
                    
                    # Check if sheet name is exactly a 4-digit number
                    if sheet_str.isdigit() and len(sheet_str) == 4:
                        year = int(sheet_str)
                        self.on_log("DEBUG", f"    -> Valid 4-digit: {year}")
                        # Validate it's a reasonable year (1900-2100)
                        if 1900 <= year <= 2100:
                            year_sheets.append(sheet_name)
                            self.on_log("INFO", f"    ✓ Accepted year sheet: {sheet_name}")
                        else:
                            self.on_log("DEBUG", f"    ✗ Year {year} outside valid range (1900-2100)")
                    else:
                        self.on_log("DEBUG", f"    ✗ Not a 4-digit number")
                
                self.on_log("INFO", f"Total year sheets found: {len(year_sheets)}")
                self.on_log("INFO", f"Year sheets: {year_sheets}")
                
                if not year_sheets:
                    messagebox.showerror("Invalid File", 
                                       "No valid year found in sheet names.\n\n" +
                                       "Sheet names must contain a 4-digit year (e.g., '2026', 'Year 2026')")
                    self.excel_file_path.set("")
                    return
                
                # Read month columns directly from ZIP (openpyxl can't load this file properly)
                self.on_log("INFO", "Reading first row to detect month columns...")
                month_columns = []
                
                try:
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        # First, load shared strings if they exist
                        shared_strings = []
                        if 'xl/sharedStrings.xml' in zip_ref.namelist():
                            with zip_ref.open('xl/sharedStrings.xml') as ss_file:
                                ss_tree = ET.parse(ss_file)
                                ss_root = ss_tree.getroot()
                                # Extract all string values
                                for elem in ss_root.iter():
                                    if elem.tag.endswith('t'):  # text element
                                        if elem.text:
                                            shared_strings.append(elem.text)
                                self.on_log("INFO", f"Loaded {len(shared_strings)} shared strings")
                        
                        # Read the first worksheet (sheet1.xml)
                        sheet_xml_path = 'xl/worksheets/sheet1.xml'
                        if sheet_xml_path in zip_ref.namelist():
                            with zip_ref.open(sheet_xml_path) as sheet_file:
                                sheet_tree = ET.parse(sheet_file)
                                sheet_root = sheet_tree.getroot()
                                
                                # Find first row
                                first_row = None
                                for elem in sheet_root.iter():
                                    if elem.tag.endswith('row'):
                                        first_row = elem
                                        break
                                
                                if first_row:
                                    self.on_log("INFO", f"Found first row")
                                    header_values = []
                                    
                                    # Get all cells (c elements) in first row
                                    for cell in first_row:
                                        if cell.tag.endswith('c'):
                                            # Check cell type - 's' means shared string
                                            cell_type = cell.attrib.get('t', '')
                                            
                                            # Get cell value
                                            v_elem = None
                                            for child in cell:
                                                if child.tag.endswith('v'):
                                                    v_elem = child
                                                    break
                                            
                                            if v_elem is not None and v_elem.text:
                                                if cell_type == 's':
                                                    # Shared string - index into shared_strings
                                                    idx = int(v_elem.text)
                                                    if 0 <= idx < len(shared_strings):
                                                        value = shared_strings[idx]
                                                    else:
                                                        value = v_elem.text
                                                else:
                                                    # Direct value
                                                    value = v_elem.text
                                                
                                                header_values.append(str(value).strip())
                                    
                                    self.on_log("INFO", f"Header values: {header_values}")
                                    
                                    # Find month columns (01-12)
                                    for val in header_values:
                                        if val.isdigit() and len(val) == 2:
                                            month_num = int(val)
                                            if 1 <= month_num <= 12:
                                                month_columns.append(val)
                                    
                                    self.on_log("INFO", f"Detected month columns: {month_columns}")
                        else:
                            self.on_log("ERROR", f"{sheet_xml_path} not found in ZIP")
                            
                except Exception as month_error:
                    self.on_log("ERROR", f"Failed to read month columns from ZIP: {month_error}")
                    import traceback
                    self.on_log("ERROR", traceback.format_exc())
                
                if not month_columns:
                    messagebox.showerror("Invalid File", 
                                       "No valid month columns found in file.\n\n" +
                                       "Expected column headers like '01', '02', '03', etc.")
                    self.excel_file_path.set("")
                    return
                
                # Update dropdowns
                self.sheet_combo['values'] = year_sheets
                
                # Convert month numbers to month names for dropdown
                month_names = ["January", "February", "March", "April", "May", "June", 
                             "July", "August", "September", "October", "November", "December"]
                available_months = [month_names[int(m)-1] for m in month_columns]
                self.month_combo['values'] = available_months
                
                # Auto-select current year and month if available
                current_year_str = str(datetime.now().year)
                if current_year_str in year_sheets:
                    self.sheet_combo.set(current_year_str)
                else:
                    self.sheet_combo.set(year_sheets[0])
                
                current_month = datetime.now().month
                current_month_str = f"{current_month:02d}"
                if current_month_str in month_columns:
                    self.month_combo.set(month_names[current_month-1])
                elif available_months:
                    self.month_combo.set(available_months[0])
                
                self.on_log("INFO", f"Selected Excel file: {file_path}")
                self.on_log("INFO", f"Available years: {', '.join(year_sheets)}")
                self.on_log("INFO", f"Available months: {', '.join(month_columns)}")
                
                # Enable Generate CSV button if month and year are selected
                if self.month_combo.get() and self.sheet_combo.get():
                    self.generate_csv_button.config(state="normal")
                
            except Exception as e:
                logger.exception("Failed to read Excel file")
                messagebox.showerror("Error", f"Failed to read Excel file:\n\n{str(e)}")
                self.on_log("ERROR", f"Failed to load file: {str(e)}")
                self.excel_file_path.set("")
    
    def _process_excel(self):
        """Process Excel file and generate receipt data."""
        try:
            # Check authentication first if portal data is needed
            if not self.web_client.is_authenticated():
                messagebox.showerror("Login Required", 
                                   "Please login before processing Excel file.\n\n" +
                                   "Contract validation requires portal access.")
                return
            
            file_path = self.excel_file_path.get()
            month = self.selected_month.get()
            sheet_name = self.selected_sheet.get()
            
            if not file_path:
                messagebox.showwarning("No File", "Please select an Excel file first.")
                return
            
            if not sheet_name:
                messagebox.showwarning("No Sheet Selected", "Please select a year tab from the dropdown.")
                return
            
            # Extract year from sheet name (handle "2026" or "Year 2026" formats)
            try:
                year = int(''.join(filter(str.isdigit, sheet_name)))
            except ValueError:
                messagebox.showerror("Invalid Sheet", f"Cannot extract year from sheet name '{sheet_name}'.\nSheet name must contain a 4-digit year.")
                return
            
            self.on_log("INFO", f"Processing Excel sheet '{sheet_name}' for month {month}")
            
            # Create a temporary compatible Excel file if openpyxl can't read the original
            temp_file = None
            try:
                # Test if openpyxl can read the file
                test_wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if not test_wb.sheetnames:
                    # File needs conversion
                    self.on_log("WARNING", "Original file not compatible with openpyxl, creating temporary copy...")
                    test_wb.close()
                    
                    # Create a new workbook and copy data from ZIP
                    import tempfile
                    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                    temp_file.close()
                    
                    self._create_compatible_excel(file_path, temp_file.name, sheet_name)
                    file_path = temp_file.name
                    self.on_log("INFO", f"Created temporary file: {temp_file.name}")
                else:
                    test_wb.close()
            except Exception as test_error:
                self.on_log("WARNING", f"File compatibility check failed: {test_error}")
            
            # Parse Excel to generate receipt data
            receipts, alerts = self.excel_processor.parse_excel(file_path, month, year, sheet_name)
            
            # Store the results
            self.current_receipts = receipts
            self.current_alerts = alerts
            
            # Count errors in alerts (rows that failed to parse)
            error_count = sum(1 for alert in alerts if 'Failed to parse' in alert.message)
            warning_count = len(alerts) - error_count
            
            # Show summary alert if there are any errors
            if error_count > 0 or warning_count > 0:
                alert_msg = f"Excel file processed with issues:\n\n"
                if error_count > 0:
                    alert_msg += f"❌ {error_count} row(s) failed to parse (skipped)\n"
                if warning_count > 0:
                    alert_msg += f"⚠️  {warning_count} warning(s)\n"
                alert_msg += f"\n✓ {len(receipts)} receipt(s) successfully processed\n\n"
                alert_msg += "Check the log for details."
                
                messagebox.showwarning("Excel Processing Issues", alert_msg)
            
            # Show alerts if any
            if alerts:
                from gui.cross_column_alert_dialog import CrossColumnAlertDialog
                dialog = CrossColumnAlertDialog(self, alerts)
                dialog.show()
            
            # Enable download button
            self.download_csv_button.config(state="normal")
            
            # Enable processing buttons after CSV generation
            self._validate_csv_file_from_receipts(receipts)
            
            self.on_log("INFO", f"✓ Excel processed: {len(receipts)} receipts ready, {len(alerts)} alerts")
            
            # Automatically validate contracts after generating receipts
            if len(receipts) > 0:
                self.on_log("INFO", "Starting automatic contract validation...")
                self._run_auto_validation()
            
            if len(receipts) == 0:
                messagebox.showwarning("No Data", 
                                     "No tenant records found in Excel file.\n\n" +
                                     "Please verify the file has data rows below the header.")
            else:
                messagebox.showinfo("Success", 
                                  f"Excel file processed successfully!\n\n" +
                                  f"Found {len(receipts)} tenant records.\n" +
                                  f"Validating contracts against portal...")
            
        except Exception as e:
            logger.exception("Failed to process Excel file")
            messagebox.showerror("Error", f"Failed to process Excel:\n\n{str(e)}")
            self.on_log("ERROR", f"Excel processing error: {str(e)}")
    
    def _download_csv(self):
        """Download the generated CSV file."""
        if not self.current_receipts:
            messagebox.showwarning("No Data", "No receipts to download. Process Excel file first.")
            return
        
        # Ask where to save CSV
        month_name = self.month_combo.get()
        sheet_name = self.selected_sheet.get()
        default_filename = f"receipts_{month_name}_{sheet_name}.csv"
        
        csv_path = filedialog.asksaveasfilename(
            title="Save CSV File",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not csv_path:
            return
        
        try:
            # Write CSV
            self._write_csv(csv_path, self.current_receipts)
            
            # Store path for potential loading
            self.csv_file_path.set(csv_path)
            
            messagebox.showinfo("Success", 
                              f"CSV downloaded successfully!\n\n{csv_path}\n\n"
                              f"{len(self.current_receipts)} receipts exported.")
            
            self.on_log("INFO", f"CSV downloaded: {csv_path}")
            
        except Exception as e:
            logger.exception("Failed to download CSV")
            messagebox.showerror("Error", f"Failed to download CSV:\n\n{str(e)}")
            self.on_log("ERROR", f"CSV download error: {str(e)}")
    
    def _browse_csv_file(self):
        """Browse for CSV file to load."""
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            self.csv_file_path.set(file_path)
            self._validate_csv_file(file_path)
            self.on_log("INFO", f"CSV file selected: {file_path}")
    
    def _load_generated_csv(self):
        """Load the generated CSV for processing."""
        if not self.current_receipts:
            messagebox.showwarning("No Data", "No receipts available. Process Excel file first.")
            return
        
        # First save to temp location if not already saved
        if not self.csv_file_path.get():
            # Auto-save to temp
            import tempfile
            month_name = self.month_combo.get()
            sheet_name = self.selected_sheet.get()
            temp_path = os.path.join(tempfile.gettempdir(), f"receipts_{month_name}_{sheet_name}.csv")
            
            try:
                self._write_csv(temp_path, self.current_receipts)
                self.csv_file_path.set(temp_path)
                self.on_log("INFO", f"Auto-saved CSV to temp: {temp_path}")
            except Exception as e:
                logger.exception("Failed to save temp CSV")
                messagebox.showerror("Error", f"Failed to prepare CSV:\n\n{str(e)}")
                return
        
        # Load CSV into handler
        csv_path = self.csv_file_path.get()
        self._validate_csv_file(csv_path)
        
        messagebox.showinfo("CSV Loaded", 
                          f"CSV loaded successfully!\n\n"
                          f"{len(self.current_receipts)} receipts ready for processing.\n\n"
                          "You can now validate contracts and process receipts.")
    
    def _write_csv(self, file_path: str, receipts: List[ReceiptData]):
        """Write receipts to CSV file."""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['contractId', 'fromDate', 'toDate', 'paymentDate', 'receiptType', 'value'])
            
            for receipt in receipts:
                writer.writerow([
                    receipt.contract_id,
                    receipt.from_date.strftime("%Y-%m-%d"),
                    receipt.to_date.strftime("%Y-%m-%d"),
                    receipt.payment_date.strftime("%Y-%m-%d"),
                    receipt.receipt_type,
                    f"{receipt.value:.2f}"
                ])
    
    def _review_csv(self):
        """Display CSV data in a table for review."""
        if not self.csv_handler.receipts:
            messagebox.showinfo("No Data", "No CSV data loaded to review.")
            return
        
        # Create a new window for the review
        review_window = tk.Toplevel(self.winfo_toplevel())
        review_window.title("CSV Review")
        review_window.geometry("900x500")
        review_window.configure(bg='#0f172a')
        
        # Record count (left-aligned)
        subtitle_frame = tk.Frame(review_window, bg='#0f172a')
        subtitle_frame.pack(fill=tk.X, padx=10, pady=(10, 10))
        
        subtitle_label = tk.Label(subtitle_frame, 
                                 text=f"Total Records: {len(self.csv_handler.receipts)}", 
                                 bg='#0f172a', fg='#94a3b8',
                                 font=('Segoe UI', 10))
        subtitle_label.pack(side=tk.LEFT)
        
        # Create frame for treeview and scrollbars
        table_frame = tk.Frame(review_window, bg='#1e293b')
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Create Treeview with scrollbars
        tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
        
        # Configure Treeview style for dark theme
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Review.Treeview",
                       background="#1e293b",
                       foreground="#e2e8f0",
                       fieldbackground="#1e293b",
                       borderwidth=0)
        style.configure("Review.Treeview.Heading",
                       background="#0f172a",
                       foreground="#3b82f6",
                       borderwidth=1,
                       relief="flat")
        style.map("Review.Treeview",
                 background=[('selected', '#3b82f6')])
        
        columns = ('Row', 'Contract ID', 'Tenant Name', 'From Date', 'To Date', 'Payment Date', 'Rent')
        tree = ttk.Treeview(table_frame, columns=columns, show='headings',
                           yscrollcommand=tree_scroll_y.set,
                           xscrollcommand=tree_scroll_x.set,
                           height=20,
                           style="Review.Treeview")
        
        tree_scroll_y.config(command=tree.yview)
        tree_scroll_x.config(command=tree.xview)
        
        # Configure columns
        tree.heading('Row', text='Row')
        tree.heading('Contract ID', text='Contract ID')
        tree.heading('Tenant Name', text='Tenant Name')
        tree.heading('From Date', text='From Date')
        tree.heading('To Date', text='To Date')
        tree.heading('Payment Date', text='Payment Date')
        tree.heading('Rent', text='Rent (€)')
        
        tree.column('Row', width=50, anchor='center')
        tree.column('Contract ID', width=100, anchor='center')
        tree.column('Tenant Name', width=200, anchor='w')
        tree.column('From Date', width=100, anchor='center')
        tree.column('To Date', width=100, anchor='center')
        tree.column('Payment Date', width=120, anchor='center')
        tree.column('Rent', width=100, anchor='e')
        
        # Add data to treeview
        for idx, receipt in enumerate(self.csv_handler.receipts, 1):
            # Get tenant name and rent value from processor's contract cache (populated after validation)
            tenant_name = 'N/A'
            rent_value = receipt.value  # Default to CSV value
            
            if hasattr(self.processor, '_contracts_data_cache'):
                # Convert contract_id to string to match cache keys
                contract_id_str = str(receipt.contract_id)
                contract_data = self.processor._contracts_data_cache.get(contract_id_str)
                if contract_data:
                    # Get tenant name(s)
                    locatarios = contract_data.get('locatarios', [])
                    if locatarios:
                        tenant_names = [t.get('nome', '') for t in locatarios]
                        tenant_name = ', '.join(filter(None, tenant_names)) or 'N/A'
                    elif contract_data.get('nomeLocatario'):
                        # Fallback to nomeLocatario field if locatarios array is empty
                        tenant_name = contract_data.get('nomeLocatario', 'N/A')
                    
                    # Get rent value from cached contract data if available
                    cached_rent = contract_data.get('valorRenda')
                    if cached_rent and (receipt.value == -1.0 or receipt.value == 0.0):
                        # Use cached rent if receipt has no value or placeholder value
                        rent_value = cached_rent
            
            tree.insert('', 'end', values=(
                idx,
                contract_id_str if hasattr(self.processor, '_contracts_data_cache') else receipt.contract_id,
                tenant_name,
                receipt.from_date,
                receipt.to_date,
                receipt.payment_date,
                f"{rent_value:.2f}" if rent_value > 0 else "N/A"
            ))
        
        # Grid layout for table components
        tree.grid(row=0, column=0, sticky='nsew')
        tree_scroll_y.grid(row=0, column=1, sticky='ns')
        tree_scroll_x.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Close button
        close_button = ThemedButton(review_window, style='secondary', text="Close", 
                                   command=review_window.destroy)
        close_button.pack(pady=(0, 10))
    
    def _validate_csv_file(self, file_path: str):
        """Validate and load CSV file."""
        # Clear cached validation when loading new CSV
        self.cached_validation_report = None
        
        success, errors = self.csv_handler.load_csv(file_path)
        
        if success:
            self.on_log("INFO", f"CSV loaded: {file_path}")
            self.on_log("INFO", f"Found {len(self.csv_handler.receipts)} receipts")
            self._update_button_states()
        else:
            # Show error details
            error_count = len(errors)
            error_msg = f"CSV file has {error_count} error(s):\n\n"
            
            # Show first 5 errors in the dialog
            display_errors = errors[:5]
            for error in display_errors:
                error_msg += f"• {error}\n"
            
            if error_count > 5:
                error_msg += f"\n... and {error_count - 5} more errors.\n"
            
            error_msg += "\nCheck the log for complete details."
            
            self.on_log("ERROR", f"Failed to load CSV: {error_count} errors found")
            for error in errors:
                self.on_log("ERROR", f"  {error}")
            
            messagebox.showerror("CSV Validation Errors", error_msg)
    
    def _validate_csv_file_from_receipts(self, receipts: List[ReceiptData]):
        """Load receipts directly into CSV handler and enable buttons."""
        # ReceiptData objects are already in the correct format for csv_handler
        # Just assign them directly
        self.csv_handler.receipts = receipts
        self._update_button_states()
    
    def _start_processing(self):
        """Start receipt processing."""
        if not self.web_client.is_authenticated():
            messagebox.showerror("Not Authenticated", "Please login first.")
            return
        
        if not self.csv_handler.receipts:
            messagebox.showwarning("No Data", "Please load a CSV file first.")
            return
        
        # Start processing in background thread
        self.stop_requested = False
        self.start_button.config(state="disabled")
        self.stop_button.config(state="normal")
        
        mode = self.mode_var.get()
        dry_run = self.dry_run_var.get()
        
        # Use csv_handler.receipts directly (they are dicts, not ReceiptData objects)
        # This matches exactly how Quick Import processes receipts
        receipts = self.csv_handler.receipts
        
        if mode == "bulk":
            self.processing_thread = threading.Thread(
                target=self._process_bulk, args=(receipts,), daemon=True
            )
        else:
            self.processing_thread = threading.Thread(
                target=self._process_step_by_step, args=(receipts,), daemon=True
            )
        
        self.processing_thread.start()
    
    def _process_bulk(self, receipts):
        """Process receipts in bulk mode."""
        self.on_log("INFO", f"Starting bulk processing of {len(receipts)} receipts")
        
        def progress_callback(current, total, message):
            if self.stop_requested:
                return
            progress = (current / total) * 100
            self.progress_var.set(progress)
            self.status_var.set(message)
            self.on_log("INFO", f"Progress: {current}/{total} - {message}")
        
        try:
            results = self.processor.process_receipts_bulk(
                receipts, 
                progress_callback, 
                validate_contracts=True,
                stop_check=lambda: self.stop_requested
            )
            if not self.stop_requested:
                self._processing_completed(results)
            else:
                self._processing_stopped()
        except Exception as e:
            if not self.stop_requested:
                self._processing_error(str(e))
    
    def _process_step_by_step(self, receipts):
        """Process receipts in step-by-step mode."""
        self.on_log("INFO", f"Starting step-by-step processing of {len(receipts)} receipts")
        
        def confirmation_callback(receipt_data, form_data):
            if self.stop_requested:
                return 'cancel'
            
            # Use thread-safe approach to show dialog on main thread
            result = [None]
            event = threading.Event()
            
            def show_dialog():
                # Import here to avoid circular import
                from gui.main_window import MainWindow
                # Create a temporary instance just to call the dialog method
                # This is not ideal but reuses the existing dialog
                result[0] = self._show_step_confirmation_dialog(receipt_data, form_data)
                event.set()
            
            # Schedule dialog on main thread
            self.after(0, show_dialog)
            
            # Wait for user response
            event.wait()
            return result[0]
        
        try:
            results = self.processor.process_receipts_step_by_step(
                receipts, 
                confirmation_callback,
                stop_check=lambda: self.stop_requested
            )
            if not self.stop_requested:
                self._processing_completed(results)
            else:
                self._processing_stopped()
        except Exception as e:
            if not self.stop_requested:
                self._processing_error(str(e))
    
    def _validate_contracts(self):
        """Show validation summary or validate CSV contract IDs against Portal das Finanças."""
        # If we have cached validation results, just show them
        if self.cached_validation_report:
            self.on_log("INFO", "Showing cached validation results...")
            self._show_validation_results(self.cached_validation_report)
            return
        
        # Otherwise, perform validation
        if not self.web_client.is_authenticated():
            messagebox.showerror("Error", "Please login first")
            return
        
        receipts = self.csv_handler.get_receipts()
        if not receipts:
            messagebox.showerror("Error", "Please load a valid CSV file first")
            return
        
        def validate():
            """Run validation in background thread."""
            try:
                self.on_log("INFO", "Starting contract validation...")
                validation_report = self.processor.validate_contracts(receipts)
                # Cache the validation report
                self.cached_validation_report = validation_report
                self.after(0, lambda: self._show_validation_results(validation_report))
            except Exception as e:
                error_msg = str(e)
                self.after(0, lambda: self._validation_error(error_msg))
        
        # Update UI
        self.validate_button.config(state="disabled")
        self.status_var.set("Validating contracts...")
        self.on_log("INFO", "Validating contract IDs against Portal das Finanças...")
        
        # Start validation in background thread
        threading.Thread(target=validate, daemon=True).start()
    
    def _run_auto_validation(self):
        """Run automatic validation after Excel processing."""
        if not self.web_client.is_authenticated():
            self.on_log("WARNING", "Not authenticated - skipping automatic validation")
            return
        
        receipts = self.csv_handler.get_receipts()
        if not receipts:
            return
        
        def validate():
            """Run validation in background thread."""
            try:
                self.on_log("INFO", "Auto-validating contracts...")
                validation_report = self.processor.validate_contracts(receipts)
                # Cache the validation report for later use
                self.cached_validation_report = validation_report
                self.after(0, lambda: self._show_validation_results(validation_report))
            except Exception as e:
                error_msg = str(e)
                self.after(0, lambda: self.on_log("ERROR", f"Auto-validation error: {error_msg}"))
        
        # Update UI
        self.status_var.set("Validating contracts...")
        
        # Start validation in background thread
        threading.Thread(target=validate, daemon=True).start()
    
    def _show_validation_results(self, validation_report: Dict[str, Any]):
        """Display contract validation results with tenant information."""
        self.validate_button.config(state="normal")
        self.status_var.set("Contract validation completed")
        
        # Log detailed results
        self.on_log("INFO", "Contract validation completed:")
        self.on_log("INFO", f"  Active portal contracts: {validation_report['portal_contracts_count']}")
        self.on_log("INFO", f"  CSV contracts: {validation_report['csv_contracts_count']}")
        self.on_log("INFO", f"  Valid matches: {len(validation_report['valid_contracts'])}")
        self.on_log("INFO", f"  Invalid contracts: {len(validation_report['invalid_contracts'])}")
        self.on_log("INFO", f"  Missing from CSV: {len(validation_report['missing_from_csv'])}")
        
        # FILTER CSV TO ONLY INCLUDE VALID CONTRACTS
        if validation_report.get('valid_contracts'):
            removed_count = self.csv_handler.filter_receipts_by_contracts(
                validation_report['valid_contracts']
            )
            if removed_count > 0:
                self.on_log("INFO", f"Filtered CSV: removed {removed_count} receipts for invalid contracts")
                self.on_log("INFO", f"Remaining receipts: {len(self.csv_handler.get_receipts())}")
        else:
            self.on_log("WARNING", "No valid contracts found - all receipts will be filtered out")
            self.csv_handler.filter_receipts_by_contracts([])
        
        # Create detailed message for popup WITH TENANT NAMES
        message_parts = []
        message_parts.append(f"✓ VALIDATION SUMMARY (Active Contracts Only):")
        message_parts.append(f"Active portal contracts: {validation_report['portal_contracts_count']}")
        message_parts.append(f"CSV contracts: {validation_report['csv_contracts_count']}")
        message_parts.append(f"Valid matches: {len(validation_report['valid_contracts'])}")
        
        # Show VALID CONTRACTS with tenant names
        if validation_report.get('valid_contracts_data'):
            message_parts.append(f"\n✅ VALID CONTRACTS (Ready for Processing):")
            for contract in validation_report['valid_contracts_data']:
                contract_id = contract.get('numero') or contract.get('referencia', 'N/A')
                tenant_name = contract.get('nomeLocatario', 'Unknown')
                rent_amount = contract.get('valorRenda', 0)
                property_addr = contract.get('imovelAlternateId', 'N/A')
                status = contract.get('estado', {}).get('label', 'Unknown')
                
                message_parts.append(f"  • {contract_id} → {tenant_name}")
                message_parts.append(f"    €{rent_amount:.2f} - {property_addr} ({status})")
        
        # Show MISSING FROM CSV (active contracts in portal but not in CSV)
        if validation_report.get('missing_from_csv_data'):
            message_parts.append(f"\n📅 ACTIVE CONTRACTS NOT TO BE ISSUED THIS MONTH:")
            for contract in validation_report['missing_from_csv_data']:
                contract_id = contract.get('numero') or contract.get('referencia', 'N/A')
                tenant_name = contract.get('nomeLocatario', 'Unknown')
                rent_amount = contract.get('valorRenda', 0)
                property_addr = contract.get('imovelAlternateId', 'N/A')
                
                message_parts.append(f"  • {contract_id} → {tenant_name}")
                message_parts.append(f"    €{rent_amount:.2f} - {property_addr}")
        
        if validation_report['validation_errors']:
            message_parts.append(f"\n VALIDATION ISSUES:")
            for error in validation_report['validation_errors']:
                message_parts.append(f"  • {error}")
    
    def _validation_error(self, error_message: str):
        """Handle validation error."""
        self.validate_button.config(state="normal")
        self.status_var.set("Contract validation failed")
        self.on_log("ERROR", f"Contract validation error: {error_message}")
        messagebox.showerror("Validation Error", f"Contract validation failed:\n{error_message}")
    
    def _export_logs(self):
        """Export the log window content to a file."""
        # Get log content from the log text widget
        log_content = self.log_text.get("1.0", tk.END).strip()
        
        if not log_content:
            messagebox.showinfo("No Logs", "No logs available to export.")
            return
        
        # Ask for file location
        default_filename = f"logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        file_path = filedialog.asksaveasfilename(
            title="Save Log File",
            defaultextension=".txt",
            initialfile=default_filename,
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Write log content to file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(log_content)
            
            self.on_log("INFO", f"Logs exported to: {file_path}")
            messagebox.showinfo("Export Successful", f"Logs exported to:\n{file_path}")
        except Exception as e:
            logger.exception("Failed to export logs")
            self.on_log("ERROR", f"Failed to export logs: {str(e)}")
            messagebox.showerror("Export Failed", f"Failed to export logs:\n{str(e)}")
    
    def _export_all_results(self):
        """Export successful receipts and error reports to separate files."""
        results = self.processor.get_results()
        if not results:
            messagebox.showinfo("Information", "No processing results to export")
            return
        
        # Check what we have to export
        successful_results = [r for r in results if r.success]
        failed_results = [r for r in results if not r.success and r.status == "Failed"]
        
        if not successful_results and not failed_results:
            messagebox.showinfo("No Data", "No results to export.")
            return
        
        success_count = 0
        error_count = 0
        success_file = None
        error_file = None
        
        # Handle successful results export with append option
        if successful_results:
            default_filename = f"successful_receipts_{datetime.now().strftime('%Y%m%d')}.csv"
            success_file = filedialog.asksaveasfilename(
                title="Save Successful Receipts (or select existing file to append)",
                defaultextension=".csv",
                initialfile=default_filename,
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
            )
            
            if success_file:
                # Check if file already exists
                file_exists = os.path.exists(success_file)
                append_mode = False
                
                if file_exists:
                    # Ask user if they want to append or overwrite
                    response = messagebox.askyesnocancel(
                        "File Exists",
                        f"The file already exists:\n{success_file}\n\n"
                        "Yes: Append successful receipts to existing file\n"
                        "No: Overwrite file with new data\n"
                        "Cancel: Cancel export"
                    )
                    
                    if response is None:  # Cancel
                        return
                    elif response:  # Yes - append
                        append_mode = True
                        self.on_log("INFO", "Appending successful receipts to existing file...")
                    else:  # No - overwrite
                        append_mode = False
                        self.on_log("INFO", "Creating new successful receipts file...")
                else:
                    self.on_log("INFO", "Creating new successful receipts file...")
                
                # Export successful receipts
                report_data = self.processor.generate_session_export_data()
                if report_data:
                    success = self.csv_handler.export_session_report(report_data, success_file, append=append_mode)
                    
                    if success:
                        success_count = len(successful_results)
                        mode_text = "appended to" if append_mode else "saved to"
                        self.on_log("INFO", f"Successful receipts {mode_text}: {success_file}")
        
        # Handle error results export
        if failed_results:
            default_error_filename = f"errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            error_file = filedialog.asksaveasfilename(
                title="Save Error Report",
                defaultextension=".csv",
                initialfile=default_error_filename,
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
            )
            
            if error_file:
                success = self.csv_handler.export_errors_report(results, error_file)
                
                if success:
                    error_count = len(failed_results)
                    self.on_log("INFO", f"Error report saved to: {error_file}")
        
        # Show summary
        if success_count > 0 or error_count > 0:
            summary_parts = []
            if success_count > 0:
                summary_parts.append(f"✅ Successful: {success_count} receipts")
                summary_parts.append(f"   File: {os.path.basename(success_file)}")
            if error_count > 0:
                summary_parts.append(f"❌ Failed: {error_count} receipts")
                summary_parts.append(f"   File: {os.path.basename(error_file)}")
            
            summary = "Export Complete!\n\n" + "\n".join(summary_parts)
            messagebox.showinfo("Export Successful", summary)
    
    def _verify_and_export_receipts(self):
        """Verify receipts in portal and export detailed report."""
        results = self.processor.get_results()
        if not results:
            messagebox.showinfo("No Results", "No receipts have been processed yet.\nPlease process receipts before verifying.")
            return
        
        # Check if authenticated
        if not self.web_client.is_authenticated():
            messagebox.showwarning("Not Authenticated", "You must be logged in to verify receipts in the portal.")
            return
        
        # Ask user for file location
        default_filename = f"verified_receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = filedialog.asksaveasfilename(
            title="Save Verified Receipts Report",
            defaultextension=".csv",
            initialfile=default_filename,
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return  # User cancelled
        
        self.on_log("INFO", "Starting receipt verification process...")
        self.status_var.set("Verifying receipts in portal...")
        
        # Run verification in a separate thread to avoid blocking the GUI
        def verify_thread():
            try:
                # Create verifier instance
                verifier = ReceiptVerifier(self.web_client)
                
                # Verify all processing results
                self.after(0, lambda: self.on_log("INFO", f"Verifying {len(results)} processed receipts..."))
                verified_receipts = verifier.verify_processing_results(results)
                
                # Export to CSV
                self.after(0, lambda: self.on_log("INFO", "Exporting verification results to CSV..."))
                success, message = verifier.export_to_csv(file_path)
                
                # Get summary stats
                stats = verifier.get_summary_stats()
                
                # Update UI on main thread
                def show_results():
                    self.status_var.set("Verification complete")
                    
                    if success:
                        summary = (
                            f"Receipt Verification Complete!\n\n"
                            f"Total Receipts: {stats['total']}\n"
                            f"✅ Verified: {stats['verified']}\n"
                            f"❌ Not Found: {stats['not_found']}\n"
                            f"⚠️  Errors: {stats['error']}\n\n"
                            f"Report saved to:\n{file_path}"
                        )
                        self.on_log("INFO", f"Verification complete: {stats['verified']}/{stats['total']} verified")
                        messagebox.showinfo("Verification Complete", summary)
                    else:
                        self.on_log("ERROR", f"Verification export failed: {message}")
                        messagebox.showerror("Export Failed", f"Failed to export verification report:\n{message}")
                
                self.after(0, show_results)
                
            except Exception as e:
                error_msg = f"Verification error: {str(e)}"
                self.after(0, lambda: self.on_log("ERROR", error_msg))
                self.after(0, lambda: messagebox.showerror("Verification Error", error_msg))
                self.after(0, lambda: self.status_var.set("Verification failed"))
        
        # Start verification thread
        thread = threading.Thread(target=verify_thread, daemon=True)
        thread.start()
    
    def _export_session_report(self):
        """
        Export session report with 2-row format per receipt.
        
        This export accumulates receipts within the same session by appending
        to the existing file if one is selected.
        """
        results = self.processor.get_results()
        if not results:
            messagebox.showinfo("No Results", "No receipts have been processed yet.\nPlease process receipts before exporting.")
            return
        
        # Check what we have to export
        successful_results = [r for r in results if r.success]
        failed_results = [r for r in results if not r.success and r.status == "Failed"]
        
        # If nothing to export at all
        if not successful_results and not failed_results:
            messagebox.showinfo("No Data", "No receipts to export.\nAll receipts were skipped.")
            return
        
        # Ask user for file location with default filename
        default_filename = f"session_report_{datetime.now().strftime('%Y%m%d')}.csv"
        file_path = filedialog.asksaveasfilename(
            title="Save Session Report",
            defaultextension=".csv",
            initialfile=default_filename,
            filetypes=[("CSV Files", "*.csv"), ("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return  # User cancelled
        
        # Determine file format based on extension
        file_ext = os.path.splitext(file_path)[1].lower()
        is_excel = file_ext == '.xlsx'
        
        # Check if file already exists
        file_exists = os.path.exists(file_path)
        append_mode = False
        
        if file_exists:
            # Ask user if they want to append or overwrite
            response = messagebox.askyesnocancel(
                "File Exists",
                f"The file already exists:\n{file_path}\n\n"
                "Yes: Append new receipts to existing file (session mode)\n"
                "No: Overwrite file with new data\n"
                "Cancel: Cancel export"
            )
            
            if response is None:  # Cancel
                return
            elif response:  # Yes - append
                append_mode = True
                self.on_log("INFO", "Appending to existing session report...")
            else:  # No - overwrite
                append_mode = False
                self.on_log("INFO", "Creating new session report...")
        else:
            self.on_log("INFO", "Creating new session report...")
        
        # Export successful receipts if any
        if successful_results:
            report_data = self.processor.generate_session_export_data()
            
            if report_data:
                # Export using appropriate method based on format
                if is_excel:
                    success = self.csv_handler.export_session_report_excel(report_data, file_path, append=append_mode)
                else:
                    success = self.csv_handler.export_session_report(report_data, file_path, append=append_mode)
                
                if success:
                    receipt_count = len(successful_results)
                    row_count = len(report_data)
                    mode_text = "appended to" if append_mode else "saved to"
                    
                    self.on_log("INFO", f"Session report {mode_text}: {file_path}")
                    self.on_log("INFO", f"Exported {receipt_count} receipts ({row_count} rows)")
                    
                    messagebox.showinfo(
                        "Export Successful",
                        f"Session report {mode_text}:\n{file_path}\n\n"
                        f"Receipts exported: {receipt_count}\n"
                        f"Total rows: {row_count}"
                    )
                else:
                    self.on_log("ERROR", "Failed to export session report")
                    messagebox.showerror("Export Failed", "Failed to export session report.\nCheck the logs for details.")
        else:
            # No successful receipts, just show info
            self.on_log("INFO", "No successful receipts to export in session report format.")
            messagebox.showinfo("No Successful Receipts", 
                              "No receipts were successfully issued.\n\n"
                              "Use the 'Export Results' button to export error details.")
        file_exists = os.path.exists(file_path)
        append_mode = False
        
        if file_exists:
            # Ask user if they want to append or overwrite
            response = messagebox.askyesnocancel(
                "File Exists",
                f"The file already exists:\n{file_path}\n\n"
                "Yes: Append new receipts to existing file (session mode)\n"
                "No: Overwrite file with new data\n"
                "Cancel: Cancel export"
            )
            
            if response is None:  # Cancel
                return
            elif response:  # Yes - append
                append_mode = True
                self.on_log("INFO", "Appending to existing session report...")
            else:  # No - overwrite
                append_mode = False
                self.on_log("INFO", "Creating new session report...")
        else:
            self.on_log("INFO", "Creating new session report...")
        
        # Generate session export data
        report_data = self.processor.generate_session_export_data()
        
        if not report_data:
            messagebox.showwarning("No Data", "No data to export")
            return
        
        # Export using appropriate method based on format
        if is_excel:
            success = self.csv_handler.export_session_report_excel(report_data, file_path, append=append_mode)
        else:
            success = self.csv_handler.export_session_report(report_data, file_path, append=append_mode)
        
        if success:
            receipt_count = len(successful_results)
            row_count = len(report_data)
            mode_text = "appended to" if append_mode else "saved to"
            
            self.on_log("INFO", f"Session report {mode_text}: {file_path}")
            self.on_log("INFO", f"Exported {receipt_count} receipts ({row_count} rows)")
            
            messagebox.showinfo(
                "Export Successful",
                f"Session report {mode_text}:\n{file_path}\n\n"
                f"Receipts exported: {receipt_count}\n"
                f"Total rows: {row_count}"
            )
        else:
            self.on_log("ERROR", "Failed to export session report")
            messagebox.showerror("Export Failed", "Failed to export session report.\nCheck the logs for details.")
    
    def _update_button_states(self):
        """Update button states based on current state."""
        has_csv = bool(self.csv_handler.receipts)
        is_authenticated = self.web_client.is_authenticated()
        
        self.start_button.config(state="normal" if (has_csv and is_authenticated) else "disabled")
        self.validate_button.config(state="normal" if (has_csv and is_authenticated) else "disabled")
        self.review_button.config(state="normal" if has_csv else "disabled")
    
    def _processing_completed(self, results: list):
        """Handle processing completion."""
        successful = sum(1 for r in results if r.success)
        skipped = sum(1 for r in results if not r.success and getattr(r, 'status', '') == 'Skipped')
        failed = sum(1 for r in results if not r.success and getattr(r, 'status', '') != 'Skipped')
        
        self.on_log("INFO", f"Processing completed. Success: {successful}, Skipped: {skipped}, Failed: {failed}")
        self.progress_var.set(100)
        
        # Create status message
        status_parts = []
        if successful > 0:
            status_parts.append(f"Success: {successful}")
        if skipped > 0:
            status_parts.append(f"Skipped: {skipped}")
        if failed > 0:
            status_parts.append(f"Failed: {failed}")
        
        status_message = f"Completed - {', '.join(status_parts)}"
        self.status_var.set(status_message)
        
        # Update UI
        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")
        
        # Show summary
        summary_parts = []
        if successful > 0:
            summary_parts.append(f"Successful: {successful}")
        if skipped > 0:
            summary_parts.append(f"Skipped: {skipped}")
        if failed > 0:
            summary_parts.append(f"Failed: {failed}")
        
        messagebox.showinfo("Processing Complete", "\n".join(summary_parts))
    
    def _processing_stopped(self):
        """Handle processing stop."""
        self.status_var.set("Stopped")
        self.on_log("WARNING", "Processing stopped by user")
        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")
        messagebox.showinfo("Stopped", "Processing was stopped.")
    
    def _processing_error(self, error_msg: str):
        """Handle processing error."""
        self.status_var.set("Error")
        self.on_log("ERROR", f"Processing error: {error_msg}")
        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")
        messagebox.showerror("Processing Error", f"An error occurred:\n\n{error_msg}")
    
    def _stop_processing(self):
        """Stop processing by setting a flag and updating UI."""
        self.on_log("INFO", "Stop requested")
        self.stop_requested = True
        self.status_var.set("Stopping...")
        
        # Update UI immediately
        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")
        
        # Reset progress
        self.progress_var.set(0)
    
    def _show_step_confirmation_dialog(self, receipt_data, form_data):
        """Show confirmation dialog for step-by-step processing."""
        # Import the main window's dialog method
        # For now, create a simple implementation
        result = ['confirm']
        
        dialog = tk.Toplevel(self)
        dialog.title("Step-by-Step Processing")
        dialog.geometry("600x400")
        dialog.configure(bg='#1e293b')
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        
        # Import ThemedButton
        from gui.themed_button import ThemedButton
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='#1e293b', padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Receipt info
        info_frame = tk.LabelFrame(main_frame, text="Receipt Information", bg='#1e293b', fg='#3b82f6',
                                  font=('Segoe UI', 9), relief='solid', borderwidth=1, padx=10, pady=10)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Contract ID
        tenant_name = form_data.get('tenant_name', '') if form_data else ''
        contract_display = str(receipt_data.contract_id)
        if tenant_name:
            contract_display = f"{receipt_data.contract_id} - {tenant_name}"
        
        tk.Label(info_frame, text=f"Contract: {contract_display}", bg='#1e293b', fg='#ffffff').pack(anchor=tk.W, pady=2)
        tk.Label(info_frame, text=f"Period: {receipt_data.from_date} to {receipt_data.to_date}", 
                bg='#1e293b', fg='#ffffff').pack(anchor=tk.W, pady=2)
        tk.Label(info_frame, text=f"Payment Date: {receipt_data.payment_date}", 
                bg='#1e293b', fg='#ffffff').pack(anchor=tk.W, pady=2)
        tk.Label(info_frame, text=f"Value: €{receipt_data.value:.2f}", 
                bg='#1e293b', fg='#ffffff').pack(anchor=tk.W, pady=2)
        
        # Question
        question_frame = tk.LabelFrame(main_frame, text="Action", bg='#1e293b', fg='#3b82f6',
                                      font=('Segoe UI', 9), relief='solid', borderwidth=1, padx=10, pady=10)
        question_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(question_frame, text="Deseja processar este recibo?", bg='#1e293b', fg='#ffffff',
                font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        
        # Buttons
        buttons_frame = tk.Frame(main_frame, bg='#1e293b')
        buttons_frame.pack(fill=tk.X)
        
        def on_confirm():
            result[0] = 'confirm'
            dialog.destroy()
        
        def on_skip():
            result[0] = 'skip'
            dialog.destroy()
        
        def on_cancel():
            result[0] = 'cancel'
            dialog.destroy()
        
        ThemedButton(buttons_frame, style='primary', text="✓ Confirm & Process", command=on_confirm).pack(side=tk.LEFT, padx=(0, 5))
        ThemedButton(buttons_frame, style='secondary', text="⏭ Skip This Receipt", command=on_skip).pack(side=tk.LEFT, padx=5)
        ThemedButton(buttons_frame, style='danger', text="✗ Cancel", command=on_cancel).pack(side=tk.LEFT, padx=(5, 0))
        
        dialog.wait_window()
        return result[0]
    
    def log(self, level: str, message: str):
        """Log message to the log text widget."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{timestamp}] {level}: {message}\n"
        
        self.log_text.insert(tk.END, log_line)
        self.log_text.see(tk.END)
        self.on_log(level, message)

