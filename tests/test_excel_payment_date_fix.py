"""
Unit tests for Excel payment date reading fix.
Verifies that payment dates are correctly extracted from Excel month column cells.
"""

import sys
import os
import pytest
import gc
import time
from datetime import datetime, date
from unittest.mock import Mock, MagicMock, patch
import tempfile

# Add the src directory to Python path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    pytest.skip("openpyxl not installed", allow_module_level=True)

from excel_preprocessor import LandlordExcelProcessor, TenantData


class TestPaymentDateReading:
    """Test that payment dates are correctly read from Excel cells."""
    
    def create_test_excel(self, payment_values):
        """
        Create a test Excel file with specified payment values.
        
        Args:
            payment_values: Dict mapping tenant names to payment date values
                           e.g. {'Tenant1': 15, 'Tenant2': datetime(2026, 1, 20)}
        
        Returns:
            Path to temporary Excel file
        """
        wb = Workbook()
        ws = wb.active
        
        # Header row
        ws['A1'] = 'Contract'
        ws['B1'] = 'Name'
        ws['C1'] = 'Rent'
        ws['D1'] = 'RentDeposit'
        ws['E1'] = 'Mes Caucao'  # Month offset column (was "PaidCurrentMonth")
        ws['F1'] = 'MonthsLate'
        ws['G1'] = 'PaidCurrentMonth'
        ws['H1'] = 'Jan'  # January column
        ws['I1'] = 'Feb'  # February column
        
        # Data rows
        row = 2
        for idx, (tenant_name, payment_value) in enumerate(payment_values.items(), start=1):
            ws[f'A{row}'] = f'12345{idx}'  # Contract
            ws[f'B{row}'] = tenant_name      # Name
            ws[f'C{row}'] = 1000.0           # Rent
            ws[f'D{row}'] = 1                # RentDeposit
            ws[f'E{row}'] = 1                # Mes Caucao (month offset)
            ws[f'F{row}'] = 0                # MonthsLate
            ws[f'G{row}'] = 'No'             # PaidCurrentMonth
            ws[f'H{row}'] = payment_value    # January payment date
            row += 1
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_path)
        wb.close()  # Close workbook to release file
        
        return temp_path
    
    def test_payment_date_from_integer(self):
        """Test reading payment date when cell contains integer day value."""
        payment_values = {
            'Tenant Day 15': 15,
            'Tenant Day 28': 28,
            'Tenant Day 5': 5
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            # Verify we got 3 receipts
            assert len(receipts) == 3, f"Expected 3 receipts, got {len(receipts)}"
            
            # Check each receipt has the correct payment date
            receipt_map = {r.contract_id: r for r in receipts}
            
            assert receipt_map['123451'].payment_date == date(2026, 1, 15), \
                f"Expected payment_date=2026-01-15, got {receipt_map['123451'].payment_date}"
            
            assert receipt_map['123452'].payment_date == date(2026, 1, 28), \
                f"Expected payment_date=2026-01-28, got {receipt_map['123452'].payment_date}"
            
            assert receipt_map['123453'].payment_date == date(2026, 1, 5), \
                f"Expected payment_date=2026-01-05, got {receipt_map['123453'].payment_date}"
            
            print("✅ Test passed: Payment dates correctly read from integer values")
            
        finally:
            # Force garbage collection before deleting
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass  # Ignore deletion errors
    
    def test_payment_date_from_datetime_object(self):
        """Test reading payment date when cell contains datetime object."""
        payment_values = {
            'Tenant Date 1': datetime(2026, 1, 12),
            'Tenant Date 2': datetime(2026, 1, 25)
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            assert len(receipts) == 2
            
            receipt_map = {r.contract_id: r for r in receipts}
            
            assert receipt_map['123451'].payment_date == date(2026, 1, 12), \
                f"Expected payment_date=2026-01-12, got {receipt_map['123451'].payment_date}"
            
            assert receipt_map['123452'].payment_date == date(2026, 1, 25), \
                f"Expected payment_date=2026-01-25, got {receipt_map['123452'].payment_date}"
            
            print("✅ Test passed: Payment dates correctly read from datetime objects")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass
    
    def test_payment_date_from_date_object(self):
        """Test reading payment date when cell contains date object."""
        payment_values = {
            'Tenant with date': date(2026, 1, 18)
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            assert len(receipts) == 1
            assert receipts[0].payment_date == date(2026, 1, 18), \
                f"Expected payment_date=2026-01-18, got {receipts[0].payment_date}"
            
            print("✅ Test passed: Payment date correctly read from date object")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass
    
    def test_payment_date_from_string_number(self):
        """Test reading payment date when cell contains string representation of day."""
        payment_values = {
            'Tenant String Day': '22'
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            assert len(receipts) == 1
            assert receipts[0].payment_date == date(2026, 1, 22), \
                f"Expected payment_date=2026-01-22, got {receipts[0].payment_date}"
            
            print("✅ Test passed: Payment date correctly read from string number")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass
    
    def test_payment_date_fallback_to_first_when_empty(self):
        """Test that receipts are skipped when payment date cell is empty."""
        payment_values = {
            'Tenant No Date': None
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            assert len(receipts) == 0, "Expected 0 receipts when payment date is empty (should be skipped)"
            
            print("✅ Test passed: Receipt skipped when payment date cell is empty")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass
    
    def test_payment_date_fallback_for_invalid_string(self):
        """Test that receipts are skipped when payment date is invalid."""
        payment_values = {
            'Tenant Invalid': 'invalid_date'
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            assert len(receipts) == 0, "Expected 0 receipts when payment date is invalid (should be skipped)"
            
            print("✅ Test passed: Receipt skipped when payment date is invalid")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass
    
    def test_payment_date_validation_out_of_range(self):
        """Test that receipts are skipped when payment day is out of range."""
        payment_values = {
            'Tenant Invalid Day': 35  # Invalid day for any month
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            assert len(receipts) == 0, "Expected 0 receipts when payment day is out of range (should be skipped)"
            
            print("✅ Test passed: Receipt skipped when payment day is out of range")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass
    
    def test_payment_date_february_validation(self):
        """Test payment date validation for February (28/29 days)."""
        wb = Workbook()
        ws = wb.active
        
        # Header
        ws['A1'] = 'Contract'
        ws['B1'] = 'Name'
        ws['C1'] = 'Rent'
        ws['D1'] = 'RentDeposit'
        ws['E1'] = 'Mes Caucao'
        ws['F1'] = 'MonthsLate'
        ws['G1'] = 'PaidCurrentMonth'
        ws['H1'] = 'Jan'
        ws['I1'] = 'Feb'
        
        # Test data - February payment
        ws['A2'] = '123456'
        ws['B2'] = 'Tenant Feb 28'
        ws['C2'] = 1000.0
        ws['D2'] = 1
        ws['E2'] = 1  # Mes Caucao
        ws['F2'] = 0
        ws['G2'] = 'No'
        ws['I2'] = 28  # Valid for February
        
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(temp_file.name, selected_month=2, selected_year=2026)
            
            assert len(receipts) == 1
            assert receipts[0].payment_date == date(2026, 2, 28), \
                f"Expected payment_date=2026-02-28, got {receipts[0].payment_date}"
            
            print("✅ Test passed: February date validation works correctly")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(temp_file.name)
            except:
                pass
    
    def test_mixed_payment_date_formats(self):
        """Test that multiple date formats can coexist, and receipts with empty dates are skipped."""
        payment_values = {
            'Tenant Integer': 10,
            'Tenant Datetime': datetime(2026, 1, 15),
            'Tenant String': '20',
            'Tenant Empty': None
        }
        
        excel_file = self.create_test_excel(payment_values)
        
        try:
            processor = LandlordExcelProcessor()
            receipts, alerts = processor.parse_excel(excel_file, selected_month=1, selected_year=2026)
            
            # 4th receipt (Tenant Empty) should be skipped
            assert len(receipts) == 3, f"Expected 3 receipts (4th with empty date skipped), got {len(receipts)}"
            
            receipt_map = {r.contract_id: r for r in receipts}
            
            assert receipt_map['123451'].payment_date == date(2026, 1, 10)
            assert receipt_map['123452'].payment_date == date(2026, 1, 15)
            assert receipt_map['123453'].payment_date == date(2026, 1, 20)
            # 123454 should not exist (skipped due to empty payment date)
            
            print("✅ Test passed: Mixed date formats handled correctly, empty date skipped")
            
        finally:
            gc.collect()
            time.sleep(0.05)
            try:
                os.unlink(excel_file)
            except:
                pass


def test_all_payment_date_scenarios():
    """Run all payment date tests and display summary."""
    if not OPENPYXL_AVAILABLE:
        print("⚠️  Skipped: openpyxl not installed")
        return
    
    print("=" * 80)
    print("EXCEL PAYMENT DATE FIX - UNIT TESTS")
    print("=" * 80)
    print()
    
    test_suite = TestPaymentDateReading()
    
    tests = [
        ("Integer Day Values", test_suite.test_payment_date_from_integer),
        ("Datetime Objects", test_suite.test_payment_date_from_datetime_object),
        ("Date Objects", test_suite.test_payment_date_from_date_object),
        ("String Numbers", test_suite.test_payment_date_from_string_number),
        ("Empty Cells (Fallback)", test_suite.test_payment_date_fallback_to_first_when_empty),
        ("Invalid Strings (Fallback)", test_suite.test_payment_date_fallback_for_invalid_string),
        ("Out of Range Days", test_suite.test_payment_date_validation_out_of_range),
        ("February Validation", test_suite.test_payment_date_february_validation),
        ("Mixed Formats", test_suite.test_mixed_payment_date_formats)
    ]
    
    passed = 0
    failed = 0
    
    for test_name, test_func in tests:
        try:
            print(f"Running: {test_name}...")
            test_func()
            passed += 1
        except Exception as e:
            print(f"❌ FAILED: {test_name}")
            print(f"   Error: {str(e)}")
            failed += 1
        print()
    
    print("=" * 80)
    print(f"TEST SUMMARY: {passed} passed, {failed} failed")
    print("=" * 80)
    print()
    
    if failed == 0:
        print("🎉 ALL TESTS PASSED!")
        print("✅ Payment dates are now correctly read from Excel cells")
        print("✅ Multiple date formats supported (datetime, date, int, string)")
        print("✅ Validation and fallback logic working correctly")
    else:
        print(f"⚠️  {failed} test(s) failed - review errors above")


if __name__ == "__main__":
    test_all_payment_date_scenarios()
