"""
Test to verify payment date is correctly parsed from dd-mm-yyyy string format.
"""

import sys
import os
import tempfile
import gc
import time
from datetime import date
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_preprocessor import LandlordExcelProcessor


def test_payment_date_dd_mm_yyyy_format():
    """
    Test that payment date is correctly parsed from dd-mm-yyyy format string.
    Example: '09-01-2026' should be parsed to day 9 for January 2026.
    """
    # Create test Excel
    wb = Workbook()
    ws = wb.active
    
    # Headers matching the user's example
    ws['A1'] = 'Contract'
    ws['B1'] = 'Nome'  # Name in Portuguese
    ws['C1'] = 'Renda'  # Rent in Portuguese
    ws['D1'] = 'Caução'  # Deposit in Portuguese
    ws['E1'] = 'Mês Caução'  # Deposit month
    ws['F1'] = 'Atraso'  # Delay
    ws['G1'] = 'PaidCurrentMonth'
    ws['H1'] = '01'  # January column
    
    # Test data - fake test data
    # 12345,Test Tenant 1,500.00,1,1,0,No,09-01-2026
    ws['A2'] = '12345'
    ws['B2'] = 'Test Tenant 1'
    ws['C2'] = 500.00  # Rent as numeric value
    ws['D2'] = 1  # Deposit
    ws['E2'] = 1  # Mes Caucao (month offset)
    ws['F2'] = 0  # Delay
    ws['G2'] = 'No'  # PaidCurrentMonth
    ws['H2'] = '09-01-2026'  # Payment date as dd-mm-yyyy string
    
    # Add more test cases with different date formats
    ws['A3'] = '67890'
    ws['B3'] = 'Test Tenant 2'
    ws['C3'] = 750.00
    ws['D3'] = 1
    ws['E3'] = 1
    ws['F3'] = 0
    ws['G3'] = 'No'
    ws['H3'] = '15-01-2026'  # 15th of January
    
    ws['A4'] = '11111'
    ws['B4'] = 'Test Tenant 3'
    ws['C4'] = 1000.00
    ws['D4'] = 1
    ws['E4'] = 1
    ws['F4'] = 0
    ws['G4'] = 'No'
    ws['H4'] = '28-01-2026'  # 28th of January
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_path = temp_file.name
    temp_file.close()
    
    wb.save(temp_path)
    wb.close()
    
    try:
        # Parse Excel for January 2026
        processor = LandlordExcelProcessor()
        receipts, alerts = processor.parse_excel(temp_path, selected_month=1, selected_year=2026)
        
        print(f"✅ Successfully parsed {len(receipts)} receipts")
        
        # Should have 3 receipts
        assert len(receipts) == 3, f"Expected 3 receipts, got {len(receipts)}"
        
        # Create a map for easier lookup
        receipt_map = {r.contract_id: r for r in receipts}
        
        # Test 1: '09-01-2026' should parse to January 9, 2026
        assert receipt_map['12345'].payment_date == date(2026, 1, 9), \
            f"Expected payment_date=2026-01-09, got {receipt_map['12345'].payment_date}"
        print(f"✅ TEST PASSED: '09-01-2026' correctly parsed to {receipt_map['12345'].payment_date}")
        
        # Test 2: '15-01-2026' should parse to January 15, 2026
        assert receipt_map['67890'].payment_date == date(2026, 1, 15), \
            f"Expected payment_date=2026-01-15, got {receipt_map['67890'].payment_date}"
        print(f"✅ TEST PASSED: '15-01-2026' correctly parsed to {receipt_map['67890'].payment_date}")
        
        # Test 3: '28-01-2026' should parse to January 28, 2026
        assert receipt_map['11111'].payment_date == date(2026, 1, 28), \
            f"Expected payment_date=2026-01-28, got {receipt_map['11111'].payment_date}"
        print(f"✅ TEST PASSED: '28-01-2026' correctly parsed to {receipt_map['11111'].payment_date}")
        
        print("\n🎉 All dd-mm-yyyy date format tests passed!")
        
    finally:
        gc.collect()
        time.sleep(0.05)
        try:
            os.unlink(temp_path)
        except:
            pass


if __name__ == '__main__':
    test_payment_date_dd_mm_yyyy_format()
