"""
Unit tests for LandlordExcelProcessor.

Tests cover Excel structure validation, tenant parsing, and receipt generation.
"""

import pytest
import sys
from pathlib import Path
from datetime import date
from openpyxl import Workbook

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

from excel_preprocessor import (
    LandlordExcelProcessor,
    TenantData,
    PaymentInfo,
    ProcessingAlert,
    ReceiptData,
    validate_excel_file
)


class TestTenantDataClass:
    """Test TenantData dataclass."""
    
    def test_tenant_data_creation(self):
        """Test creating TenantData object."""
        tenant = TenantData(
            contract_number='12345',
            name='John Doe',
            rent=500.00,
            rent_deposit=1,
            deposit_month_offset=1,
            months_late=0,
            paid_current_month=False,
            row_number=2
        )
        
        assert tenant.contract_number == '12345'
        assert tenant.name == 'John Doe'
        assert tenant.rent == 500.00
        assert tenant.rent_deposit == 1
        assert tenant.deposit_month_offset == 1
        assert tenant.months_late == 0
        assert tenant.paid_current_month is False
        assert tenant.row_number == 2


class TestPaymentInfoClass:
    """Test PaymentInfo dataclass."""
    
    def test_payment_info_creation(self):
        """Test creating PaymentInfo object."""
        tenant = TenantData('12345', 'John', 500.0, 1, 1, 0, False, 2)
        payment = PaymentInfo(
            tenant=tenant,
            payment_date=date(2026, 6, 5),
            payment_column='Jun',
            payment_day=5
        )
        
        assert payment.tenant == tenant
        assert payment.payment_date == date(2026, 6, 5)
        assert payment.payment_column == 'Jun'
        assert payment.payment_day == 5


class TestProcessingAlertClass:
    """Test ProcessingAlert dataclass."""
    
    def test_alert_creation(self):
        """Test creating ProcessingAlert object."""
        alert = ProcessingAlert(
            contract_number='12345',
            payment_date=date(2026, 6, 15),
            payment_column='Jun',
            expected_column='May',
            rent_period_from=date(2026, 5, 1),
            rent_period_to=date(2026, 5, 31),
            reason='Late payment - paid in June for May rent'
        )
        
        assert alert.contract_number == '12345'
        assert alert.payment_column == 'Jun'
        assert alert.expected_column == 'May'


class TestReceiptDataClass:
    """Test ReceiptData dataclass."""
    
    def test_receipt_data_creation(self):
        """Test creating ReceiptData object."""
        receipt = ReceiptData(
            contract_id='12345',
            from_date=date(2026, 6, 1),
            to_date=date(2026, 6, 30),
            payment_date=date(2026, 6, 5),
            receipt_type='rent',
            value=500.00
        )
        
        assert receipt.contract_id == '12345'
        assert receipt.from_date == date(2026, 6, 1)
        assert receipt.to_date == date(2026, 6, 30)
        assert receipt.payment_date == date(2026, 6, 5)
        assert receipt.receipt_type == 'rent'
        assert receipt.value == 500.00


class TestLandlordExcelProcessorInitialization:
    """Test processor initialization."""
    
    def test_processor_initialization(self):
        """Test processor initializes with empty state."""
        processor = LandlordExcelProcessor()
        
        assert processor.validation_errors == []
        assert processor.processing_alerts == []


class TestColumnMapping:
    """Test column mapping functionality."""
    
    def test_build_column_map_standard_headers(self):
        """Test column mapping with standard header names."""
        processor = LandlordExcelProcessor()
        
        headers = ['Contract', 'Name', 'Rent', 'RentDeposit', 'MonthsLate', 
                  'PaidCurrentMonth', 'Jan', 'Feb', 'Mar']
        
        col_map = processor._build_column_map(headers)
        
        assert col_map['contract'] == 0
        assert col_map['name'] == 1
        assert col_map['rent'] == 2
        assert col_map['rent_deposit'] == 3
        assert col_map['months_late'] == 4
        assert col_map['paid_current_month'] == 5
        assert col_map['month_columns'][1] == 6  # Jan
        assert col_map['month_columns'][2] == 7  # Feb
        assert col_map['month_columns'][3] == 8  # Mar
    
    def test_build_column_map_case_insensitive(self):
        """Test column mapping handles different cases."""
        processor = LandlordExcelProcessor()
        
        headers = ['contract', 'NAME', 'Rent', 'rentdeposit', 'MONTHSLATE',
                  'paidcurrentmonth']
        
        col_map = processor._build_column_map(headers)
        
        assert col_map['contract'] == 0
        assert col_map['name'] == 1
        assert col_map['rent'] == 2
        assert col_map['rent_deposit'] == 3
        assert col_map['months_late'] == 4
        assert col_map['paid_current_month'] == 5
    
    def test_build_column_map_with_spaces(self):
        """Test column mapping handles headers with spaces."""
        processor = LandlordExcelProcessor()
        
        headers = ['Contract Number', 'Tenant Name', 'Monthly Rent', 
                  'Rent Deposit', 'Months Late', 'Paid Current Month']
        
        col_map = processor._build_column_map(headers)
        
        assert col_map['contract'] == 0
        assert col_map['name'] == 1
        assert col_map['rent'] == 2
        assert col_map['rent_deposit'] == 3
        assert col_map['months_late'] == 4
        assert col_map['paid_current_month'] == 5


class TestLandlordSeparatorDetection:
    """Test landlord separator row detection."""
    
    def test_detects_landlord_prefix(self):
        """Test detects rows starting with LANDLORD."""
        processor = LandlordExcelProcessor()
        
        row = ['[LANDLORD: Property Owner A]', None, None, None]
        assert processor._is_landlord_separator_row(row) is True
    
    def test_detects_property_owner(self):
        """Test detects PROPERTY OWNER text."""
        processor = LandlordExcelProcessor()
        
        row = ['PROPERTY OWNER: John Smith', None, None]
        assert processor._is_landlord_separator_row(row) is True
    
    def test_detects_owner_colon(self):
        """Test detects OWNER: pattern."""
        processor = LandlordExcelProcessor()
        
        row = ['Owner: Jane Doe', None, None]
        assert processor._is_landlord_separator_row(row) is True
    
    def test_normal_row_not_separator(self):
        """Test normal tenant row is not detected as separator."""
        processor = LandlordExcelProcessor()
        
        row = ['12345', 'John Tenant', 500, 1, 0, 'No']
        assert processor._is_landlord_separator_row(row) is False
    
    def test_empty_row_not_separator(self):
        """Test empty row is not detected as separator."""
        processor = LandlordExcelProcessor()
        
        row = [None, None, None]
        assert processor._is_landlord_separator_row(row) is False


class TestYesNoParser:
    """Test Yes/No value parsing."""
    
    def test_parse_yes_variations(self):
        """Test various YES representations."""
        processor = LandlordExcelProcessor()
        
        assert processor._parse_yes_no('Yes') is True
        assert processor._parse_yes_no('YES') is True
        assert processor._parse_yes_no('yes') is True
        assert processor._parse_yes_no('Y') is True
        assert processor._parse_yes_no('y') is True
        assert processor._parse_yes_no('TRUE') is True
        assert processor._parse_yes_no('1') is True
    
    def test_parse_no_variations(self):
        """Test various NO representations."""
        processor = LandlordExcelProcessor()
        
        assert processor._parse_yes_no('No') is False
        assert processor._parse_yes_no('NO') is False
        assert processor._parse_yes_no('no') is False
        assert processor._parse_yes_no('N') is False
        assert processor._parse_yes_no('FALSE') is False
        assert processor._parse_yes_no('0') is False
        assert processor._parse_yes_no(None) is False
        assert processor._parse_yes_no('') is False


class TestTenantRowParsing:
    """Test tenant row parsing."""
    
    def test_parse_valid_tenant_row(self):
        """Test parsing a valid tenant row."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['12345', 'John Doe', 500.0, 1, 1, 0, 'No']
        
        tenant = processor._parse_tenant_row(row_values, col_map, 2)
        
        assert tenant.contract_number == '12345'
        assert tenant.name == 'John Doe'
        assert tenant.rent == 500.0
        assert tenant.rent_deposit == 1
        assert tenant.deposit_month_offset == 1
        assert tenant.months_late == 0
        assert tenant.paid_current_month is False
        assert tenant.row_number == 2
    
    def test_parse_tenant_with_paid_current_month(self):
        """Test parsing tenant with PaidCurrentMonth = Yes."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['12346', 'Jane Smith', 600.0, 0, 1, 0, 'Yes']
        
        tenant = processor._parse_tenant_row(row_values, col_map, 3)
        
        assert tenant.contract_number == '12346'
        assert tenant.paid_current_month is True
    
    def test_parse_tenant_with_months_late(self):
        """Test parsing tenant with late payments."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['11111', 'Paul Late', 600.0, 1, 1, 2, 'No']
        
        tenant = processor._parse_tenant_row(row_values, col_map, 4)
        
        assert tenant.months_late == 2
    
    def test_parse_missing_contract_raises_error(self):
        """Test missing contract number raises ValueError."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['', 'Unknown', 600.0, 1, 1, 0, 'No']
        
        with pytest.raises(ValueError, match="Missing contract number"):
            processor._parse_tenant_row(row_values, col_map, 2)
    
    def test_parse_invalid_rent_raises_error(self):
        """Test invalid rent amount raises ValueError."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['12345', 'John Doe', 'invalid', 1, 1, 0, 'No']
        
        with pytest.raises(ValueError, match="Invalid rent amount"):
            processor._parse_tenant_row(row_values, col_map, 2)
    
    def test_parse_negative_rent_raises_error(self):
        """Test negative rent raises ValueError."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['12345', 'John Doe', -500.0, 1, 1, 0, 'No']
        
        with pytest.raises(ValueError, match="Invalid rent amount"):
            processor._parse_tenant_row(row_values, col_map, 2)
    
    def test_parse_invalid_rent_deposit_raises_error(self):
        """Test invalid RentDeposit raises ValueError."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['12345', 'John Doe', 500.0, 'invalid', 1, 0, 'No']
        
        with pytest.raises(ValueError, match="Invalid RentDeposit"):
            processor._parse_tenant_row(row_values, col_map, 2)
    
    def test_parse_invalid_months_late_raises_error(self):
        """Test invalid MonthsLate raises ValueError."""
        processor = LandlordExcelProcessor()
        
        col_map = {
            'contract': 0,
            'name': 1,
            'rent': 2,
            'rent_deposit': 3,
            'deposit_month_offset': 4,
            'months_late': 5,
            'paid_current_month': 6,
            'month_columns': {}
        }
        
        row_values = ['12345', 'John Doe', 500.0, 1, 1, 'late', 'No']
        
        with pytest.raises(ValueError, match="Invalid MonthsLate"):
            processor._parse_tenant_row(row_values, col_map, 2)


class TestExpectedColumnCalculation:
    """Test expected column calculation for cross-column detection."""
    
    def test_expected_column_normal_payment(self):
        """Test expected column for normal on-time payment."""
        processor = LandlordExcelProcessor()
        
        # Payment on June 5, RentDeposit=1, MonthsLate=0
        # Expected: Paying for July (June + 1 month)
        payment_date = date(2026, 6, 5)
        
        expected = processor._calculate_expected_column(
            payment_date, rent_deposit=1, months_late=0, paid_current_month=False
        )
        
        assert expected == 'Jul'
    
    def test_expected_column_paid_current_month(self):
        """Test expected column with PaidCurrentMonth flag."""
        processor = LandlordExcelProcessor()
        
        # Payment on June 5, PaidCurrentMonth=True
        # Expected: June (current month)
        payment_date = date(2026, 6, 5)
        
        expected = processor._calculate_expected_column(
            payment_date, rent_deposit=1, months_late=0, paid_current_month=True
        )
        
        assert expected == 'Jun'
    
    def test_expected_column_late_payment(self):
        """Test expected column for late payment."""
        processor = LandlordExcelProcessor()
        
        # Payment on June 15, RentDeposit=1, MonthsLate=2
        # Expected: Paying for May (June + 1 - 2 = May)
        payment_date = date(2026, 6, 15)
        
        expected = processor._calculate_expected_column(
            payment_date, rent_deposit=1, months_late=2, paid_current_month=False
        )
        
        assert expected == 'May'


class TestExcelStructureValidation:
    """Test Excel structure validation."""
    
    def test_validate_empty_file(self):
        """Test validation fails for empty file."""
        processor = LandlordExcelProcessor()
        
        # Create minimal workbook with only header
        wb = Workbook()
        ws = wb.active
        ws.append(['Contract', 'Name', 'Rent'])
        
        processor._validate_excel_structure(ws)
        
        assert len(processor.validation_errors) > 0
        assert "empty" in processor.validation_errors[0].lower()
    
    def test_validate_missing_required_columns(self):
        """Test validation fails when required columns missing."""
        processor = LandlordExcelProcessor()
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Contract', 'Name'])  # Missing Rent and other required columns
        ws.append(['12345', 'John'])
        
        processor._validate_excel_structure(ws)
        
        assert len(processor.validation_errors) > 0
        errors_text = ' '.join(processor.validation_errors).lower()
        assert 'rent' in errors_text or 'missing' in errors_text
    
    def test_validate_no_month_columns(self):
        """Test validation fails when no month columns present."""
        processor = LandlordExcelProcessor()
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Contract', 'Name', 'Rent', 'RentDeposit', 'MonthsLate', 
                  'PaidCurrentMonth'])  # No month columns
        ws.append(['12345', 'John', 500, 1, 0, 'No'])
        
        processor._validate_excel_structure(ws)
        
        assert len(processor.validation_errors) > 0
        assert any('month' in err.lower() for err in processor.validation_errors)
    
    def test_validate_valid_structure(self):
        """Test validation passes for valid structure."""
        processor = LandlordExcelProcessor()
        
        wb = Workbook()
        ws = wb.active
        ws.append(['Contract', 'Name', 'Rent', 'RentDeposit', 'MonthsLate',
                  'PaidCurrentMonth', 'Jan', 'Feb', 'Mar'])
        ws.append(['12345', 'John', 500, 1, 0, 'No', 5, None, None])
        
        processor._validate_excel_structure(ws)
        
        assert len(processor.validation_errors) == 0


# Note: Full integration tests with actual Excel files would be in 
# test_smart_import_integration.py
