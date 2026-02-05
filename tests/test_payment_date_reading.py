"""
Test to verify payment date is read correctly from month column.
"""

import sys
import os
import tempfile
import gc
import time
from datetime import date
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_preprocessor import LandlordExcelProcessor


def test_payment_date_from_month_column():
    """
    Test that payment date day is read from the selected month column.
    """
    # Create test Excel
    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws['A1'] = 'Contract'
    ws['B1'] = 'Name'
    ws['C1'] = 'Rent'
    ws['D1'] = 'RentDeposit'
    ws['E1'] = 'Mes Caucao'
    ws['F1'] = 'MonthsLate'
    ws['G1'] = 'PaidCurrentMonth'
    ws['H1'] = '01'  # January column
    ws['I1'] = '02'  # February column
    
    # Test data - payment day 25 in January
    ws['A2'] = '12345'
    ws['B2'] = 'Test Tenant'
    ws['C2'] = 1000.0
    ws['D2'] = 1
    ws['E2'] = 1
    ws['F2'] = 0
    ws['G2'] = 'No'
    ws['H2'] = 25  # Payment day in January
    ws['I2'] = 10  # Different day in February
    
    # Save and close
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_path = temp_file.name
    temp_file.close()
    
    wb.save(temp_path)
    wb.close()
    
    try:
        # Parse Excel for January
        processor = LandlordExcelProcessor()
        receipts, alerts = processor.parse_excel(temp_path, selected_month=1, selected_year=2026)
        
        assert len(receipts) == 1, f"Expected 1 receipt, got {len(receipts)}"
        
        receipt = receipts[0]
        
        # Payment date should be January 25, 2026
        assert receipt.payment_date == date(2026, 1, 25), \
            f"Expected payment_date=2026-01-25, got {receipt.payment_date}"
        
        print("✅ TEST PASSED: Payment date correctly read as Jan 25 from '01' column")
        
        # Now test February
        receipts2, alerts2 = processor.parse_excel(temp_path, selected_month=2, selected_year=2026)
        
        assert len(receipts2) == 1, f"Expected 1 receipt, got {len(receipts2)}"
        
        receipt2 = receipts2[0]
        
        # Payment date should be February 10, 2026
        assert receipt2.payment_date == date(2026, 2, 10), \
            f"Expected payment_date=2026-02-10, got {receipt2.payment_date}"
        
        print("✅ TEST PASSED: Payment date correctly read as Feb 10 from '02' column")
        
    finally:
        gc.collect()
        time.sleep(0.05)
        try:
            os.unlink(temp_path)
        except:
            pass


if __name__ == '__main__':
    test_payment_date_from_month_column()
    print("\n🎉 Payment date reading test passed!")
