"""
Test edge cases for dd-mm-yyyy date format parsing.
"""

import sys
import os
import tempfile
import gc
import time
from datetime import date
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_preprocessor import LandlordExcelProcessor


def test_edge_cases():
    """Test edge cases in date format parsing."""
    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws['A1'] = 'Contract'
    ws['B1'] = 'Name'
    ws['C1'] = 'Rent'
    ws['D1'] = 'RentDeposit'  # Changed from 'Deposit'
    ws['E1'] = 'Mes Caucao'
    ws['F1'] = 'MonthsLate'
    ws['G1'] = 'PaidCurrentMonth'
    ws['H1'] = '01'  # January
    
    # Test cases
    test_cases = [
        ('123001', 'Space Before', '  09-01-2026', 9),  # Leading spaces
        ('123002', 'Space After', '09-01-2026  ', 9),   # Trailing spaces
        ('123003', 'Single Digit', '5-01-2026', 5),      # Single digit day
        ('123004', 'Just Number', '15', 15),             # Just a number (old format)
        ('123005', 'Full Date', '31-01-2026', 31),       # End of month
    ]
    
    row = 2
    for contract, name, date_value, expected_day in test_cases:
        ws[f'A{row}'] = contract
        ws[f'B{row}'] = name
        ws[f'C{row}'] = 1000.0
        ws[f'D{row}'] = 1
        ws[f'E{row}'] = 1
        ws[f'F{row}'] = 0
        ws[f'G{row}'] = 'No'
        ws[f'H{row}'] = date_value
        row += 1
    
    # Save
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_path = temp_file.name
    temp_file.close()
    
    wb.save(temp_path)
    wb.close()
    
    try:
        processor = LandlordExcelProcessor()
        receipts, alerts = processor.parse_excel(temp_path, selected_month=1, selected_year=2026)
        
        print(f"✅ Parsed {len(receipts)} receipts")
        
        receipt_map = {r.contract_id: r for r in receipts}
        
        # Verify each test case
        for contract, name, date_value, expected_day in test_cases:
            actual_day = receipt_map[contract].payment_date.day
            if actual_day == expected_day:
                print(f"✅ {name}: '{date_value}' -> day {actual_day}")
            else:
                print(f"❌ {name}: Expected day {expected_day}, got {actual_day}")
                raise AssertionError(f"Date parsing failed for {name}")
        
        print("\n🎉 All edge case tests passed!")
        
    finally:
        gc.collect()
        time.sleep(0.05)
        try:
            os.unlink(temp_path)
        except:
            pass


if __name__ == '__main__':
    test_edge_cases()
