"""
Test to verify fromDate and toDate read from Mes Caucao column (not Caucao column).
"""

import sys
import os
import tempfile
import gc
import time
from datetime import date
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_preprocessor import LandlordExcelProcessor


def test_from_to_dates_use_mes_caucao_column():
    """
    Test that fromDate and toDate calculations use the Mes Caucao column
    (not the Caucao/RentDeposit column).
    """
    # Create test Excel with both Caucao and Mes Caucao columns
    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws['A1'] = 'Contract'
    ws['B1'] = 'Name'
    ws['C1'] = 'Rent'
    ws['D1'] = 'RentDeposit'  # Caucao - this should NOT be used for date calculation
    ws['E1'] = 'Mes Caucao'   # This SHOULD be used for date calculation
    ws['F1'] = 'MonthsLate'
    ws['G1'] = 'PaidCurrentMonth'
    ws['H1'] = 'Jan'
    
    # Test data:
    # - RentDeposit (Caucao) = 2 (if used incorrectly, would calculate March)
    # - Mes Caucao = 1 (correct value, should calculate February)
    ws['A2'] = '12345'
    ws['B2'] = 'Test Tenant'
    ws['C2'] = 1000.0
    ws['D2'] = 2  # RentDeposit = 2 (WRONG column for date calc)
    ws['E2'] = 1  # Mes Caucao = 1 (CORRECT column for date calc)
    ws['F2'] = 0
    ws['G2'] = 'No'
    ws['H2'] = 15  # Payment day
    
    # Save and close
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_path = temp_file.name
    temp_file.close()
    
    wb.save(temp_path)
    wb.close()
    
    try:
        # Parse Excel (payment made in January 2026)
        processor = LandlordExcelProcessor()
        receipts, alerts = processor.parse_excel(temp_path, selected_month=1, selected_year=2026)
        
        # Verify
        assert len(receipts) == 1, f"Expected 1 receipt, got {len(receipts)}"
        
        receipt = receipts[0]
        
        # Payment date should be January 15, 2026
        assert receipt.payment_date == date(2026, 1, 15), \
            f"Expected payment_date=2026-01-15, got {receipt.payment_date}"
        
        # fromDate should be February 1, 2026 (Jan + Mes Caucao value of 1)
        # NOT March 1 (which would be if using RentDeposit value of 2)
        assert receipt.from_date == date(2026, 2, 1), \
            f"Expected from_date=2026-02-01 (using Mes Caucao=1), got {receipt.from_date}. " \
            f"If this is 2026-03-01, it means the code is using RentDeposit (Caucao) instead of Mes Caucao!"
        
        # toDate should be February 28, 2026 (last day of February)
        assert receipt.to_date == date(2026, 2, 28), \
            f"Expected to_date=2026-02-28, got {receipt.to_date}"
        
        print("✅ TEST PASSED: fromDate and toDate correctly read from Mes Caucao column")
        print(f"   - RentDeposit (Caucao) = 2 (not used for dates)")
        print(f"   - Mes Caucao = 1 (correctly used for dates)")
        print(f"   - from_date = {receipt.from_date} (Feb 1, not Mar 1)")
        print(f"   - to_date = {receipt.to_date} (Feb 28)")
        
    finally:
        gc.collect()
        time.sleep(0.05)
        try:
            os.unlink(temp_path)
        except:
            pass


if __name__ == '__main__':
    test_from_to_dates_use_mes_caucao_column()
    print("\n🎉 All from/to date tests passed!")
